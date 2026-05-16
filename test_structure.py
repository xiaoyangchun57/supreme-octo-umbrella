import os
import openpyxl

# 设置路径
structure_path = r'E:\杂七杂八\trea项目\xiang\dist\模板\对应表.xlsx'

print(f"对应表路径: {structure_path}")
print(f"对应表存在: {os.path.exists(structure_path)}")

if os.path.exists(structure_path):
    wb = openpyxl.load_workbook(structure_path, data_only=True)
    
    print(f"\n工作表数量: {len(wb.sheetnames)}")
    print(f"工作表名称: {wb.sheetnames}")
    
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"\n=== Sheet{idx}: {sheet_name} ===")
        
        # 检查前5行前5列
        for row in range(1, 6):
            row_data = []
            for col in range(1, 6):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(f"列{col}: '{cell_value}'")
            print(f"第{row}行: {', '.join(row_data)}")
    
    wb.close()
