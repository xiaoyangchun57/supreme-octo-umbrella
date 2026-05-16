import os
import math
import random
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from ..utils import read_csv_file, log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR as OUTPUT_DIR, DEFAULT_TEMPLATE_DIR as TEMPLATE_DIR

class CsvToReportConverter:
    def __init__(self, output_dir=None, template_dir=None):
        self.results = {
            'success': [],
            'failed': [],
            'total': 0
        }
        self.output_dir = output_dir if output_dir else OUTPUT_DIR
        self.template_dir = template_dir if template_dir else TEMPLATE_DIR
    
    def adjust_csv_points(self, csv_data):
        """调整CSV点数据：查找ZJ/YJ，计算投影，排序点"""
        if not csv_data or len(csv_data) < 2:
            return None, None, None
        
        zj_row = None
        yj_row = None
        
        for i, row in enumerate(csv_data):
            if len(row) >= 5:
                marker = str(row[4]).strip().upper()
                if marker == 'ZJ':
                    zj_row = i
                elif marker == 'YJ':
                    yj_row = i
        
        if zj_row is None or yj_row is None:
            return None, None, None
        
        zj_y = float(csv_data[zj_row][1]) if csv_data[zj_row][1] else 0.0
        zj_x = float(csv_data[zj_row][2]) if csv_data[zj_row][2] else 0.0
        yj_y = float(csv_data[yj_row][1]) if csv_data[yj_row][1] else 0.0
        yj_x = float(csv_data[yj_row][2]) if csv_data[yj_row][2] else 0.0
        
        line_vector_x = yj_x - zj_x
        line_vector_y = yj_y - zj_y
        vector_length_sq = line_vector_x ** 2 + line_vector_y ** 2
        
        if vector_length_sq <= 0.001:
            return None, None, None
        
        proj_points = []
        for i, row in enumerate(csv_data):
            if len(row) >= 3:
                try:
                    x = float(row[2]) if row[2] else 0.0
                    y = float(row[1]) if row[1] else 0.0
                    dot_product = (x - zj_x) * line_vector_x + (y - zj_y) * line_vector_y
                    t = dot_product / vector_length_sq
                    proj_points.append((i, t, row))
                except:
                    proj_points.append((i, 0, row))
        
        proj_points.sort(key=lambda x: x[1])
        
        sorted_data = [p[2] for p in proj_points]
        
        new_zj_row = None
        new_yj_row = None
        for i, row in enumerate(sorted_data):
            if len(row) >= 5:
                marker = str(row[4]).strip().upper()
                if marker == 'ZJ':
                    new_zj_row = i
                elif marker == 'YJ':
                    new_yj_row = i
        
        if new_zj_row != 0 and new_zj_row is not None:
            zj_data = sorted_data.pop(new_zj_row)
            sorted_data.insert(0, zj_data)
        
        if new_yj_row is not None:
            if new_yj_row > 0:
                idx = next((i for i, r in enumerate(sorted_data) if len(r)>=5 and str(r[4]).strip().upper()=='YJ'), None)
                if idx is not None and idx != len(sorted_data)-1:
                    yj_data = sorted_data.pop(idx)
                    sorted_data.append(yj_data)
        
        last_row = len(sorted_data)
        if last_row >= 2:
            zj_x = float(sorted_data[0][2]) if sorted_data[0][2] else 0.0
            zj_y = float(sorted_data[0][1]) if sorted_data[0][1] else 0.0
            yj_x = float(sorted_data[-1][2]) if sorted_data[-1][2] else 0.0
            yj_y = float(sorted_data[-1][1]) if sorted_data[-1][1] else 0.0
            
            line_vector_x = yj_x - zj_x
            line_vector_y = yj_y - zj_y
            vector_length_sq = line_vector_x ** 2 + line_vector_y ** 2
            
            for i in range(1, last_row - 1):
                if len(sorted_data[i]) >= 3:
                    try:
                        x = float(sorted_data[i][2]) if sorted_data[i][2] else 0.0
                        y = float(sorted_data[i][1]) if sorted_data[i][1] else 0.0
                        dot_product = (x - zj_x) * line_vector_x + (y - zj_y) * line_vector_y
                        t = dot_product / vector_length_sq
                        sorted_data[i][2] = str(zj_x + t * line_vector_x)
                        sorted_data[i][1] = str(zj_y + t * line_vector_y)
                    except:
                        pass
        
        return sorted_data, (zj_x, zj_y), (yj_x, yj_y)
    
    def is_levee_marker(self, marker):
        """判断是否为堤坝标记"""
        marker = str(marker).strip().upper()
        return marker in ['KS', 'ZKS', 'YKS', 'DD', 'ZDD', 'YDD']
    
    def gauss_proj_inv(self, x, y, central_meridian=117.0):
        """高斯投影反算：将平面坐标转换为经纬度"""
        a = 6378137.0
        f = 1.0 / 298.257223563
        e2 = 2 * f - f * f
        e12 = e2 / (1 - e2)
        
        central_meridian_rad = central_meridian * math.pi / 180.0
        y0 = y - 500000
        
        Bf_rad = x / (a * (1 - e2 / 4 - 3 * e2 * e2 / 64 - 5 * e2 * e2 * e2 / 256))
        
        for _ in range(5):
            sinBf = math.sin(Bf_rad)
            cosBf = math.cos(Bf_rad)
            
            M = a * ((1 - e2 / 4 - 3 * e2 * e2 / 64 - 5 * e2 * e2 * e2 / 256) * Bf_rad -
                    (3 * e2 / 8 + 3 * e2 * e2 / 32 + 45 * e2 * e2 * e2 / 1024) * math.sin(2 * Bf_rad) +
                    (15 * e2 * e2 / 256 + 45 * e2 * e2 * e2 / 1024) * math.sin(4 * Bf_rad) -
                    (35 * e2 * e2 * e2 / 3072) * math.sin(6 * Bf_rad))
            
            Bf0_rad = Bf_rad
            Bf_rad = (x - M) / (a * (1 - e2)) + Bf_rad
            if abs(Bf_rad - Bf0_rad) < 0.0000000001:
                break
        
        sinBf = math.sin(Bf_rad)
        cosBf = math.cos(Bf_rad)
        tf = sinBf / cosBf
        tf2 = tf * tf
        nf2 = e12 * cosBf * cosBf
        Nf = a / math.sqrt(1 - e2 * sinBf * sinBf)
        Mf = a * (1 - e2) / ((1 - e2 * sinBf * sinBf) ** 1.5)
        
        delta_l_rad = y0 / Nf
        
        B_rad = Bf_rad - (y0 * y0 * tf) / (2 * Mf * Nf) + \
                (y0 ** 4 * tf) / (24 * Mf * Nf ** 3) * (5 + 3 * tf2 + nf2 - 9 * nf2 * tf2) - \
                (y0 ** 6 * tf) / (720 * Mf * Nf ** 5) * (61 + 90 * tf2 + 45 * tf2 * tf2)
        
        L_rad = delta_l_rad / cosBf - \
                (1 + 2 * tf2 + nf2) * (delta_l_rad ** 3) / (6 * cosBf) + \
                (5 + 28 * tf2 + 24 * tf2 * tf2 + 6 * nf2 + 8 * nf2 * tf2) * (delta_l_rad ** 5) / (120 * cosBf)
        
        B = B_rad * 180.0 / math.pi
        L = central_meridian + L_rad * 180.0 / math.pi
        
        return B, L
    
    def calculate_azimuth(self, zj_lon, zj_lat, yj_lon, yj_lat):
        """计算方位角"""
        if zj_lon == yj_lon and zj_lat == yj_lat:
            return 0.0
        
        zj_lat_rad = math.radians(zj_lat)
        zj_lon_rad = math.radians(zj_lon)
        yj_lat_rad = math.radians(yj_lat)
        yj_lon_rad = math.radians(yj_lon)
        
        delta_lon = yj_lon_rad - zj_lon_rad
        
        y = math.cos(yj_lat_rad) * math.sin(delta_lon)
        x = math.cos(zj_lat_rad) * math.sin(yj_lat_rad) - \
            math.sin(zj_lat_rad) * math.cos(yj_lat_rad) * math.cos(delta_lon)
        
        azimuth_rad = math.atan2(y, x)
        azimuth_deg = math.degrees(azimuth_rad)
        
        while azimuth_deg < 0:
            azimuth_deg += 360
        while azimuth_deg >= 360:
            azimuth_deg -= 360
        
        return azimuth_deg
    
    def mark_deepest_point(self, ws, start_row, end_row):
        """标记最深点"""
        min_elevation = float('inf')
        min_row = 0
        
        for i in range(start_row, end_row + 1):
            cell_value = ws.cell(row=i, column=4).value
            if isinstance(cell_value, (int, float)):
                if cell_value < min_elevation:
                    min_elevation = cell_value
                    min_row = i
        
        if min_row > 0:
            ws.cell(row=min_row, column=2, value='深泓点')
    
    def convert_cross_section(self, csv_file, progress_callback=None):
        """转换横断面CSV到成果表"""
        try:
            csv_data = read_csv_file(csv_file)
            if not csv_data:
                raise ValueError("CSV文件为空")
            
            template_path = os.path.join(self.template_dir, '横断面成果表模板.xlsx')
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"模板文件不存在: {template_path}")
            
            sorted_data, zj_point, yj_point = self.adjust_csv_points(csv_data)
            if sorted_data is None:
                raise ValueError("无法找到ZJ或YJ点")
            
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            levee_points = []
            zj_x, zj_y = zj_point
            
            for i, row in enumerate(sorted_data):
                if len(row) >= 5:
                    try:
                        x = float(row[2]) if row[2] else 0.0
                        y = float(row[1]) if row[1] else 0.0
                        distance = math.sqrt((x - zj_x) ** 2 + (y - zj_y) ** 2)
                        marker = str(row[4]).strip().upper()
                        if self.is_levee_marker(marker):
                            levee_points.append((i, distance, marker))
                    except:
                        pass
            
            min_levee_row = None
            max_levee_row = None
            if levee_points:
                min_levee = min(levee_points, key=lambda x: x[1])
                max_levee = max(levee_points, key=lambda x: x[1])
                min_levee_row = min_levee[0]
                max_levee_row = max_levee[0]
            
            xlsx_row = 13
            zj_lon = 0.0
            zj_lat = 0.0
            yj_lon = 0.0
            yj_lat = 0.0
            
            for i, row in enumerate(sorted_data):
                if len(row) < 5:
                    continue
                
                try:
                    x = float(row[2]) if row[2] else 0.0
                    y = float(row[1]) if row[1] else 0.0
                    distance = math.sqrt((x - zj_x) ** 2 + (y - zj_y) ** 2)
                    elevation = float(row[3]) if row[3] else 0.0
                    marker = str(row[4]).strip().upper()
                    code = str(row[5]).strip() if len(row) > 5 else ''
                except:
                    continue
                
                final_marker = marker
                if self.is_levee_marker(marker):
                    if i == min_levee_row:
                        final_marker = '左堤顶'
                    elif i == max_levee_row:
                        final_marker = '右堤顶'
                
                try:
                    lat, lon = self.gauss_proj_inv(y, x)
                except:
                    lat, lon = 0.0, 0.0
                
                if marker == 'ZJ':
                    zj_lon = lon
                    zj_lat = lat
                    ws.cell(row=8, column=2, value=round(lon, 6))
                    ws.cell(row=8, column=5, value=round(lat, 6))
                    ws.cell(row=7, column=5, value=round(elevation, 3))
                elif marker == 'ZZS':
                    ws.cell(row=10, column=2, value=round(elevation, 3))
                elif marker == 'LSH':
                    ws.cell(row=9, column=5, value=round(elevation, 3))
                elif marker == 'YJ':
                    yj_lon = lon
                    yj_lat = lat
                
                ws.cell(row=xlsx_row, column=1, value=xlsx_row - 12)
                ws.cell(row=xlsx_row, column=2, value=final_marker)
                ws.cell(row=xlsx_row, column=3, value=round(distance, 1))
                ws.cell(row=xlsx_row, column=4, value=round(elevation, 3))
                ws.cell(row=xlsx_row, column=5, value=round(lon, 6))
                ws.cell(row=xlsx_row, column=6, value=round(lat, 6))
                ws.cell(row=xlsx_row, column=7, value=code)
                
                xlsx_row += 1
            
            if zj_lon != 0 and yj_lon != 0:
                azimuth = self.calculate_azimuth(zj_lon, zj_lat, yj_lon, yj_lat)
                ws.cell(row=9, column=2, value=round(azimuth, 4))
            
            self.mark_deepest_point(ws, 14, xlsx_row - 1)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            for row in range(13, xlsx_row):
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    cell.border = thin_border
            
            for row in range(13, xlsx_row):
                ws.cell(row=row, column=3).number_format = '0.0'
                ws.cell(row=row, column=4).number_format = '0.000'
                ws.cell(row=row, column=5).number_format = '0.000000'
                ws.cell(row=row, column=6).number_format = '0.000000'
            
            ws.cell(row=9, column=2).number_format = '0.0000'
            ws.cell(row=10, column=2).number_format = '0.000'
            ws.cell(row=7, column=5).number_format = '0.000'
            ws.cell(row=9, column=5).number_format = '0.000'
            ws.cell(row=8, column=2).number_format = '0.000000'
            ws.cell(row=8, column=5).number_format = '0.000000'
            
            file_name = os.path.basename(csv_file)
            output_name = f"{os.path.splitext(file_name)[0]}.xlsx"
            output_path = os.path.join(self.output_dir, output_name)
            
            wb.save(output_path)
            self.results['success'].append(output_name)
            log_message(f"成功转换横断面: {file_name}")
            
            if progress_callback:
                progress_callback(f"已转换: {file_name}")
            
            return True
        except Exception as e:
            file_name = os.path.basename(csv_file)
            self.results['failed'].append((file_name, str(e)))
            log_error(f"转换横断面失败: {file_name}", e)
            return False
    
    def atan2(self, y, x):
        """自定义Atan2函数"""
        if x > 0:
            return math.atan(y / x)
        elif x < 0:
            if y >= 0:
                return math.atan(y / x) + math.pi
            else:
                return math.atan(y / x) - math.pi
        else:
            if y > 0:
                return math.pi / 2
            elif y < 0:
                return -math.pi / 2
            else:
                return 0
    
    def ensure_e_greater_than_d_and_decreasing(self, data_array, e_fixed):
        """确保E列值大于D列值且单调递减"""
        n = len(data_array)
        
        for i in range(n):
            if e_fixed[i] is None or e_fixed[i] <= data_array[i][3]:
                e_fixed[i] = data_array[i][3] + 0.001 + random.random() * 0.5
        
        for i in range(1, n):
            if e_fixed[i] >= e_fixed[i-1]:
                decline_amount = 0.001 + random.random() * 0.5
                e_fixed[i] = e_fixed[i-1] - decline_amount
            
            if e_fixed[i] <= data_array[i][3]:
                e_fixed[i] = data_array[i][3] + 0.001 + random.random() * 0.5
        
        return e_fixed
    
    def convert_longitudinal_section(self, csv_file, progress_callback=None):
        """转换纵断面CSV到成果表"""
        try:
            csv_data = read_csv_file(csv_file)
            if not csv_data:
                raise ValueError("CSV文件为空")
            
            template_path = os.path.join(self.template_dir, '纵断面模板.xlsx')
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"模板文件不存在: {template_path}")
            
            wb = openpyxl.load_workbook(template_path)
            
            try:
                ws = wb['沟道纵断面测量成果表']
            except:
                ws = wb.active
            
            if len(csv_data) < 2:
                raise ValueError("CSV数据行数不足")
            
            central_meridian = 117.0
            
            first_lat, first_lon = self.gauss_proj_inv(float(csv_data[0][1]) if csv_data[0][1] else 0.0, 
                                                      float(csv_data[0][2]) if csv_data[0][2] else 0.0,
                                                      central_meridian)
            
            ws.cell(row=5, column=4, value=round(first_lon, 6))
            ws.cell(row=6, column=2, value=round(first_lat, 6))
            ws.cell(row=6, column=4, value=float(csv_data[0][3]) if csv_data[0][3] else 0.0)
            
            last_row = len(csv_data)
            if last_row < 2:
                raise ValueError("数据不足")
            
            try:
                d2_value = float(csv_data[1][3]) if csv_data[1][3] else 0.0
            except:
                d2_value = 0.0
            try:
                d_last_value = float(csv_data[-1][3]) if csv_data[-1][3] else 0.0
            except:
                d_last_value = 0.0
            
            data_array = []
            for i in range(1, last_row):
                row = csv_data[i]
                station = str(row[0]).strip() if row[0] else ''
                try:
                    y = float(row[1]) if row[1] else 0.0
                except:
                    y = 0.0
                try:
                    x = float(row[2]) if row[2] else 0.0
                except:
                    x = 0.0
                try:
                    elevation = float(row[3]) if row[3] else 0.0
                except:
                    elevation = 0.0
                try:
                    e_value = float(row[4]) if len(row) > 4 and row[4] else None
                except:
                    e_value = None
                data_array.append([station, y, x, elevation, e_value])
            
            if d_last_value > d2_value:
                data_array = data_array[::-1]
            
            directions = [0.0]
            for i in range(1, len(data_array)):
                delta_y = data_array[i][1] - data_array[i-1][1]
                delta_x = data_array[i][2] - data_array[i-1][2]
                
                direction = math.degrees(self.atan2(delta_y, delta_x))
                if direction < 0:
                    direction += 360
                directions.append(direction)
            
            e_fixed = [None] * len(data_array)
            for i in range(len(data_array)):
                if data_array[i][4] is not None:
                    e_fixed[i] = data_array[i][4]
            
            i = 0
            while i < len(data_array):
                if e_fixed[i] is None:
                    missing_start = i
                    while i < len(data_array) and e_fixed[i] is None:
                        i += 1
                    missing_end = i - 1
                    
                    if missing_start > 0:
                        prev_valid = e_fixed[missing_start - 1]
                    else:
                        prev_valid = e_fixed[missing_end + 1] if missing_end + 1 < len(data_array) else 0.0
                    
                    if missing_end < len(data_array) - 1:
                        next_valid = e_fixed[missing_end + 1]
                    else:
                        next_valid = e_fixed[missing_start - 1] if missing_start > 0 else 0.0
                    
                    diff = next_valid - prev_valid
                    avg_step = diff / ((missing_end - missing_start + 1) + 1)
                    
                    current_value = prev_valid
                    for j in range(missing_start, missing_end + 1):
                        current_value = current_value + avg_step
                        rand_offset = (random.random() - 0.5) * 1.0
                        if rand_offset > 0.501:
                            rand_offset = 0.501
                        if rand_offset < -0.501:
                            rand_offset = -0.501
                        current_value = current_value + rand_offset
                        
                        if current_value > next_valid:
                            current_value = next_valid - 0.01
                        if current_value < prev_valid:
                            current_value = prev_valid + 0.01
                        
                        e_fixed[j] = current_value
                else:
                    i += 1
            
            e_fixed = self.ensure_e_greater_than_d_and_decreasing(data_array, e_fixed)
            
            start_row = 11
            
            for row in range(start_row, ws.max_row + 1):
                for col in range(1, 9):
                    ws.cell(row=row, column=col).value = None
            
            for i in range(len(data_array)):
                ws.cell(row=start_row + i, column=1).value = i + 1
                
                if i == 0:
                    ws.cell(row=start_row + i, column=2).value = "起点"
                else:
                    ws.cell(row=start_row + i, column=2).value = "测点" + str(i)
                
                if i == 0:
                    ws.cell(row=start_row + i, column=3).value = 0
                else:
                    point_distance = math.sqrt(
                        (data_array[i][2] - data_array[i-1][2]) ** 2 +
                        (data_array[i][1] - data_array[i-1][1]) ** 2
                    )
                    ws.cell(row=start_row + i, column=3).value = round(point_distance, 1)
                
                ws.cell(row=start_row + i, column=4).value = round(directions[i], 1)
                ws.cell(row=start_row + i, column=5).value = data_array[i][3]
                ws.cell(row=start_row + i, column=6).value = e_fixed[i]
                
                lat, lon = self.gauss_proj_inv(data_array[i][1], data_array[i][2], central_meridian)
                ws.cell(row=start_row + i, column=7).value = round(lon, 6)
                ws.cell(row=start_row + i, column=8).value = round(lat, 6)
            
            last_x_row = start_row + len(data_array) - 1
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            for row in range(start_row, last_x_row + 1):
                for col in range(1, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    cell.border = thin_border
            
            for row in range(start_row, last_x_row + 1):
                ws.cell(row=row, column=3).number_format = '0.0'
                ws.cell(row=row, column=4).number_format = '0'
                ws.cell(row=row, column=5).number_format = '0.000'
                ws.cell(row=row, column=6).number_format = '0.000'
                ws.cell(row=row, column=7).number_format = '0.000000'
                ws.cell(row=row, column=8).number_format = '0.000000'
            
            file_name = os.path.basename(csv_file)
            output_name = f"{os.path.splitext(file_name)[0]}.xlsx"
            output_path = os.path.join(self.output_dir, output_name)
            
            wb.save(output_path)
            self.results['success'].append(output_name)
            log_message(f"成功转换纵断面: {file_name}")
            
            if progress_callback:
                progress_callback(f"已转换: {file_name}")
            
            return True
        except Exception as e:
            file_name = os.path.basename(csv_file)
            self.results['failed'].append((file_name, str(e)))
            log_error(f"转换纵断面失败: {file_name}", e)
            return False
    
    def convert_bridge_section(self, csv_file, progress_callback=None):
        """转换桥断面CSV到成果表"""
        return self.convert_cross_section(csv_file, progress_callback)
    
    def convert_storage_section(self, csv_file, progress_callback=None):
        """转换库容断面CSV到成果表"""
        try:
            csv_data = read_csv_file(csv_file)
            if not csv_data:
                raise ValueError("CSV文件为空")
            
            template_path = os.path.join(self.template_dir, '横断面成果表模板.xlsx')
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"模板文件不存在: {template_path}")
            
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            zj_x, zj_y = 0.0, 0.0
            yj_x, yj_y = 0.0, 0.0
            zj_found = False
            yj_found = False
            
            levee_points = []
            
            for row in csv_data:
                if len(row) >= 5:
                    marker = str(row[4]).strip().upper()
                    if marker == 'ZJ':
                        try:
                            zj_x = float(row[2]) if row[2] else 0.0
                            zj_y = float(row[1]) if row[1] else 0.0
                            zj_elevation = float(row[3]) if row[3] else 0.0
                            zj_found = True
                            
                            lat, lon = self.gauss_proj_inv(zj_y, zj_x)
                            ws.cell(row=8, column=2, value=round(lon, 6))
                            ws.cell(row=8, column=5, value=round(lat, 6))
                            ws.cell(row=7, column=5, value=round(zj_elevation, 3))
                        except:
                            pass
                    elif marker == 'YJ':
                        try:
                            yj_x = float(row[2]) if row[2] else 0.0
                            yj_y = float(row[1]) if row[1] else 0.0
                            yj_found = True
                        except:
                            pass
                    elif marker in ['ZDD', 'ZKS', 'KS', 'DD']:
                        try:
                            x = float(row[2]) if row[2] else 0.0
                            y = float(row[1]) if row[1] else 0.0
                            distance = math.sqrt((x - zj_x) ** 2 + (y - zj_y) ** 2)
                            levee_points.append((distance, marker, 'left'))
                        except:
                            pass
                    elif marker in ['YDD', 'YKS']:
                        try:
                            x = float(row[2]) if row[2] else 0.0
                            y = float(row[1]) if row[1] else 0.0
                            distance = math.sqrt((x - zj_x) ** 2 + (y - zj_y) ** 2)
                            levee_points.append((distance, marker, 'right'))
                        except:
                            pass
            
            if zj_found and yj_found:
                zj_lat, zj_lon = self.gauss_proj_inv(zj_y, zj_x)
                yj_lat, yj_lon = self.gauss_proj_inv(yj_y, yj_x)
                azimuth = self.calculate_azimuth(zj_lon, zj_lat, yj_lon, yj_lat)
                ws.cell(row=9, column=2, value=round(azimuth, 4))
            
            min_levee_marker = None
            max_levee_marker = None
            if levee_points:
                min_levee = min(levee_points, key=lambda x: x[0])
                max_levee = max(levee_points, key=lambda x: x[0])
                if min_levee[2] == 'left':
                    min_levee_marker = min_levee[1]
                if max_levee[2] == 'right':
                    max_levee_marker = max_levee[1]
            
            data_rows = []
            xlsx_row = 13
            for i, row in enumerate(csv_data):
                if len(row) >= 4:
                    try:
                        x = float(row[2]) if row[2] else 0.0
                        y = float(row[1]) if row[1] else 0.0
                        elevation = float(row[3]) if row[3] else 0.0
                        marker = str(row[4]).strip().upper() if len(row) > 4 else str(i + 1)
                        
                        if zj_found:
                            distance = math.sqrt((x - zj_x) ** 2 + (y - zj_y) ** 2)
                        else:
                            distance = x
                        
                        lat, lon = self.gauss_proj_inv(y, x)
                        
                        final_marker = marker
                        if marker in ['ZDD', 'ZKS', 'KS', 'DD'] and marker == min_levee_marker:
                            final_marker = '左堤顶'
                        elif marker in ['YDD', 'YKS'] and marker == max_levee_marker:
                            final_marker = '右堤顶'
                        
                        data_rows.append({
                            'row': xlsx_row,
                            'elevation': elevation,
                            'marker': final_marker,
                            'x': x, 'y': y, 'distance': distance,
                            'lat': lat, 'lon': lon
                        })
                        
                        xlsx_row += 1
                    except:
                        continue
            
            if data_rows:
                min_elev_row = min(data_rows, key=lambda r: r['elevation'])
                min_elev_row['marker'] = '深泓点'
            
            for row_data in data_rows:
                ws.cell(row=row_data['row'], column=1, value=row_data['row'] - 12)
                ws.cell(row=row_data['row'], column=2, value=row_data['marker'])
                ws.cell(row=row_data['row'], column=3, value=round(row_data['distance'], 1))
                ws.cell(row=row_data['row'], column=4, value=round(row_data['elevation'], 3))
                ws.cell(row=row_data['row'], column=5, value=round(row_data['lon'], 6))
                ws.cell(row=row_data['row'], column=6, value=round(row_data['lat'], 6))
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            for row in range(13, xlsx_row):
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    cell.border = thin_border
            
            file_name = os.path.basename(csv_file)
            output_name = f"{os.path.splitext(file_name)[0]}.xlsx"
            output_path = os.path.join(self.output_dir, output_name)
            
            wb.save(output_path)
            self.results['success'].append(output_name)
            log_message(f"成功转换库容断面: {file_name}")
            
            if progress_callback:
                progress_callback(f"已转换: {file_name}")
            
            return True
        except Exception as e:
            file_name = os.path.basename(csv_file)
            self.results['failed'].append((file_name, str(e)))
            log_error(f"转换库容断面失败: {file_name}", e)
            return False
    
    def process_all(self, csv_files, progress_callback=None):
        """批量处理所有CSV文件"""
        self.results = {'success': [], 'failed': [], 'total': len(csv_files)}
        
        for csv_file in csv_files:
            file_name = os.path.basename(csv_file)
            prefix = file_name[0].upper()
            
            if prefix in ['B', 'G', 'J']:
                self.convert_cross_section(csv_file, progress_callback)
            elif prefix == 'Z':
                self.convert_longitudinal_section(csv_file, progress_callback)
            elif prefix == 'Q':
                self.convert_bridge_section(csv_file, progress_callback)
            elif prefix == 'K':
                self.convert_storage_section(csv_file, progress_callback)
        
        if progress_callback and self.results['failed']:
            progress_callback(f"\n===== 失败详情 =====")
            for item in self.results['failed']:
                progress_callback(f"失败: {item[0]} - {item[1]}")
        
        return self.results