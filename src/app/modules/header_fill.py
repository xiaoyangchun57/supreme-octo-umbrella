import os
import openpyxl
import gc
import time
from ..utils import log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR as OUTPUT_DIR, DEFAULT_TEMPLATE_DIR as TEMPLATE_DIR

class HeaderFiller:
    def __init__(self, output_dir=None, template_dir=None):
        self.results = {
            'success': [],
            'failed': [],
            'total': 0
        }
        self.dict1 = {}  # Sheet1数据（横断面）
        self.dict2 = {}  # Sheet2数据（纵断面）
        self.output_dir = output_dir if output_dir else OUTPUT_DIR
        self.template_dir = template_dir if template_dir else TEMPLATE_DIR
    
    def load_mapping_table(self, sheet_type):
        """加载对应表数据
        
        对应表结构：
        - 第1列：短名称（用于查找）
        - 第11列：长名称（用于重命名）
        """
        try:
            table_path = os.path.join(self.template_dir, '对应表.xlsx')
            if not os.path.exists(table_path):
                table_path = os.path.join(self.template_dir, '对应表.xlsm')
                if not os.path.exists(table_path):
                    raise FileNotFoundError(f"对应表文件不存在: {table_path}")
            
            wb = openpyxl.load_workbook(table_path, data_only=True)
            if sheet_type == 1:
                ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.worksheets[0]
                current_dict = self.dict1
            else:
                ws = wb['Sheet2'] if 'Sheet2' in wb.sheetnames else wb.worksheets[1]
                current_dict = self.dict2
            
            current_dict.clear()
            
            last_row = ws.max_row
            for i in range(2, last_row + 1):
                key_cell = ws.cell(row=i, column=1)
                if key_cell.value is not None:
                    key = str(key_cell.value).strip()
                    arr = [''] * 12
                    for j in range(1, 13):
                        if j <= ws.max_column:
                            cell_val = ws.cell(row=i, column=j).value
                            arr[j-1] = str(cell_val).strip() if cell_val else ''
                    
                    if key not in current_dict:
                        current_dict[key] = []
                    current_dict[key].append(arr)
            
            wb.close()
            log_message(f"成功加载Sheet{sheet_type}数据，共 {len(current_dict)} 个关键字")
            return True
        except Exception as e:
            log_error(f"加载对应表Sheet{sheet_type}失败", e)
            return False
    
    def get_base_name(self, full_name):
        """获取文件名（不含扩展名）"""
        return os.path.splitext(full_name)[0]
    
    def is_valid_excel_file(self, file_name):
        """判断是否为有效Excel文件"""
        lower_name = file_name.lower()
        return (lower_name.endswith('.xlsx') or lower_name.endswith('.xlsm')) and \
               lower_name != '对应表.xlsx' and \
               lower_name != '对应表.xlsm' and \
               not lower_name.startswith('~$')
    
    def update_cells_stage1(self, ws, arr_data):
        """横断面表头填写"""
        ws['B2'] = arr_data[1]
        ws['E2'] = arr_data[2]
        ws['B3'] = arr_data[3]
        ws['E3'] = arr_data[4]
        ws['B4'] = arr_data[5]
        ws['B5'] = arr_data[6]
        ws['E4'] = arr_data[7]
        ws['E5'] = arr_data[8]
        ws['E6'] = arr_data[9]
    
    def update_cells_stage2(self, ws, arr_data):
        """纵断面表头填写"""
        ws['B2'] = arr_data[1]
        ws['D4'] = arr_data[2]
        ws['B3'] = arr_data[3]
        ws['D3'] = arr_data[4]
        ws['B4'] = arr_data[5]
    
    def process_all(self, report_files=None, progress_callback=None):
        """批量处理所有文件
        
        逻辑说明：
        1. 用文件名（短名称）在对应表第1列查找
        2. 找到后填写表头数据
        3. 从第11列获取长名称进行重命名
        
        注意：短名称和长名称之间没有字符串关系，唯一的关系是在同一行
        """
        self.results = {'success': [], 'failed': [], 'total': 0}
        
        try:
            if report_files is None:
                report_files = [os.path.join(self.output_dir, f) for f in os.listdir(self.output_dir) if f.endswith('.xlsx')]
            
            if not self.load_mapping_table(1):
                log_error("加载Sheet1失败", "无法继续")
                return self.results
            
            if not self.load_mapping_table(2):
                log_error("加载Sheet2失败", "无法继续")
                return self.results
            
            rename_mapping = {}
            
            total_files = len(report_files)
            completed_count = 0
            
            for file_path in report_files:
                if not os.path.isfile(file_path):
                    continue
                    
                file_name = os.path.basename(file_path)
                
                if not self.is_valid_excel_file(file_name):
                    continue
                
                base_name = self.get_base_name(file_name)
                success = False
                new_name = None
                reason = None
                
                log_message(f"处理文件: {file_name}")
                log_message(f"  短名称: {base_name}")
                log_message(f"  在Sheet1中存在: {base_name in self.dict1}")
                log_message(f"  在Sheet2中存在: {base_name in self.dict2}")
                
                if base_name in self.dict1:
                    data_list = self.dict1[base_name]
                    if data_list and len(data_list) > 0:
                        arr_data = data_list[0]
                        log_message(f"  在Sheet1找到数据，长名称: {arr_data[10]}")
                        try:
                            wb = None
                            try:
                                wb = openpyxl.load_workbook(file_path)
                                ws = wb.active
                                self.update_cells_stage1(ws, arr_data)
                                wb.save(file_path)
                            finally:
                                if wb:
                                    wb.close()
                                    wb = None
                            
                            new_name = arr_data[10] + '.xlsx' if arr_data[10] else file_name
                            success = True
                            log_message(f"  横断面处理成功: {file_name} -> {new_name}")
                        except Exception as e:
                            reason = f"横断面处理失败: {str(e)}"
                            log_error(f"  错误: {reason}", e)
                
                if not success and base_name in self.dict2:
                    data_list = self.dict2[base_name]
                    if data_list and len(data_list) > 0:
                        arr_data = data_list[0]
                        log_message(f"  在Sheet2找到数据，长名称: {arr_data[7]}")
                        try:
                            wb = None
                            try:
                                wb = openpyxl.load_workbook(file_path)
                                ws = wb.active
                                self.update_cells_stage2(ws, arr_data)
                                wb.save(file_path)
                            finally:
                                if wb:
                                    wb.close()
                                    wb = None
                            
                            new_name = arr_data[7] + '.xlsx' if arr_data[7] else file_name
                            success = True
                            log_message(f"  纵断面处理成功: {file_name} -> {new_name}")
                        except Exception as e:
                            reason = f"纵断面处理失败: {str(e)}"
                            log_error(f"  错误: {reason}", e)
                
                if success:
                    self.results['success'].append(file_name)
                    if new_name and new_name != file_name:
                        rename_mapping[file_name] = new_name
                else:
                    if reason is None:
                        reason = f"在对应表中未找到: {base_name}"
                    self.results['failed'].append((file_name, reason))
                    log_message(f"  跳过: {file_name} - {reason}")
                
                completed_count += 1
                if progress_callback:
                    progress_callback(f"处理: {file_name}")
                    progress_callback(f"已完成 {completed_count}/{total_files}")
            
            log_message(f"\n等待文件句柄释放...")
            gc.collect()
            time.sleep(0.5)
            
            log_message(f"\n准备重命名 {len(rename_mapping)} 个文件...")
            for old_name, new_name in rename_mapping.items():
                old_path = os.path.join(self.output_dir, old_name)
                new_path = os.path.join(self.output_dir, new_name)
                log_message(f"  准备重命名: {old_name} -> {new_name}")
                log_message(f"  原文件存在: {os.path.exists(old_path)}")
                log_message(f"  新文件存在: {os.path.exists(new_path)}")
                try:
                    if os.path.exists(old_path) and not os.path.exists(new_path):
                        os.rename(old_path, new_path)
                        log_message(f"  ✓ 重命名成功: {old_name} -> {new_name}")
                    elif os.path.exists(new_path):
                        log_message(f"  ✗ 目标文件已存在，跳过重命名: {new_name}")
                        # 将重命名失败记录到结果中
                        if (old_name, "目标文件已存在") not in self.results['failed']:
                            self.results['failed'].append((old_name, "目标文件已存在"))
                    else:
                        log_message(f"  ✗ 原文件不存在: {old_path}")
                        if (old_name, "原文件不存在") not in self.results['failed']:
                            self.results['failed'].append((old_name, "原文件不存在"))
                except Exception as e:
                    log_error(f"  ✗ 重命名失败: {old_name} -> {new_name}", e)
                    if (old_name, f"重命名失败: {str(e)}") not in self.results['failed']:
                        self.results['failed'].append((old_name, f"重命名失败: {str(e)}"))
            
            self.results['total'] = len(self.results['success'])
            log_message(f"表头填写完成，成功处理 {self.results['total']} 个文件")
            
            if progress_callback and self.results['failed']:
                progress_callback(f"\n===== 失败详情 =====")
                for item in self.results['failed']:
                    progress_callback(f"失败: {item[0]} - {item[1]}")
            
            return self.results
        except Exception as e:
            log_error("处理过程发生错误", e)
            return self.results
