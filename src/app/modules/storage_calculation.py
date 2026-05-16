import os
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from ..utils import log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR as OUTPUT_DIR, DEFAULT_TEMPLATE_DIR as TEMPLATE_DIR
import re

class StorageCalculator:
    def __init__(self, output_dir=None, template_dir=None):
        self.results = {
            'success': [],
            'failed': [],
            'total': 0
        }
        self.output_dir = output_dir if output_dir else OUTPUT_DIR
        self.template_dir = template_dir if template_dir else TEMPLATE_DIR
        self.log_text = ""
        self.progress_callback = None
        self.ref_data = None
        self.stop_flag = False
        
        self.storage_output_dir = os.path.join(self.output_dir, '库容计算')
        self.section_area_template_path = os.path.join(self.template_dir, '断面面积计算模板.xlsx')
        self.storage_data_template_path = os.path.join(self.template_dir, '库容数据模板.xlsx')
    
    def stop(self):
        """停止计算"""
        self.stop_flag = True
        self._log_action("  收到停止信号，正在停止计算...")
    
    def _check_stop(self):
        """检查是否需要停止计算"""
        if self.stop_flag:
            self._log_action("  计算已停止")
            return True
        return False
    
    def _log_action(self, message):
        """记录日志"""
        self.log_text += message + "\n"
        log_message(message)
        if self.progress_callback:
            self.progress_callback(message)
    
    def _load_ref_data(self):
        """加载对应表数据（Sheet3 - 高程差值）"""
        try:
            table_path = os.path.join(self.template_dir, '对应表.xlsx')
            if not os.path.exists(table_path):
                table_path = os.path.join(self.template_dir, '对应表.xlsm')
                if not os.path.exists(table_path):
                    table_path = os.path.join(self.output_dir, '对应表.xlsx')
            
            if not os.path.exists(table_path):
                self._log_action("警告: 对应表不存在，无法进行高程转换")
                return False
            
            wb = openpyxl.load_workbook(table_path, data_only=True)
            if 'Sheet3' in wb.sheetnames:
                ws = wb['Sheet3']
            else:
                ws = wb.active
            
            self.ref_data = []
            for row in ws.iter_rows(values_only=True):
                self.ref_data.append(list(row))
            
            self._log_action("成功加载对应表数据")
            return True
        except Exception as e:
            self._log_action(f"加载对应表数据失败: {str(e)}")
            return False
    
    def _find_subtract_value(self, section_name):
        """根据断面编号查找对应的高程差值"""
        if not self.ref_data or len(self.ref_data) < 2:
            return None
        
        search_name = str(section_name).strip()
        
        for row in self.ref_data[1:]:
            for idx, cell_value in enumerate(row):
                if cell_value and str(cell_value).strip() == search_name:
                    if idx < len(self.ref_data[0]):
                        subtract_value = self.ref_data[0][idx]
                        if isinstance(subtract_value, (int, float)):
                            return subtract_value
        return None
    
    def _extract_bridge_and_elevation(self, cell_value):
        """从单元格值中提取桥名和库容顶点高程"""
        if not cell_value:
            return None, None
        
        text = str(cell_value).strip()
        
        elevation_match = re.search(r'(\d+\.?\d*)$', text)
        if elevation_match:
            elevation = float(elevation_match.group(1))
            bridge_name = text[:elevation_match.start()].strip()
            return bridge_name, elevation
        
        return text, None
    
    def _calculate_section_area(self, points, vertex_elevation):
        """计算断面面积（考虑天然河道凹凸不平）"""
        if not points or len(points) < 2:
            return 0.0
        
        sorted_points = sorted(points, key=lambda x: x[0])
        
        area = 0.0
        
        for i in range(len(sorted_points) - 1):
            x1, y1 = sorted_points[i]
            x2, y2 = sorted_points[i + 1]
            
            if y1 is None or y2 is None:
                continue
            
            depth1 = max(0, vertex_elevation - y1)
            depth2 = max(0, vertex_elevation - y2)
            
            if depth1 > 0 or depth2 > 0:
                width = abs(x2 - x1)
                avg_depth = (depth1 + depth2) / 2
                area += width * avg_depth
        
        return round(area, 4)
    
    def _find_deepest_point(self, points):
        """查找深泓点（最低点）"""
        if not points:
            return None
        
        deepest = None
        min_elevation = float('inf')
        
        for x, y in points:
            if y is not None and y < min_elevation:
                min_elevation = y
                deepest = (x, y)
        
        return deepest
    
    def _calculate_distance(self, point1, point2):
        """计算两点之间的平面距离"""
        if not point1 or not point2:
            return 0.0
        
        x1, y1 = point1
        x2, y2 = point2
        
        return ((x1 - x2) ** 2 + (y1 - y2) ** 2) ** 0.5
    
    def _calculate_distance_by_latlon(self, lon1, lat1, lon2, lat2):
        """使用经纬度计算两点之间的距离（Haversine公式）"""
        if lon1 is None or lat1 is None or lon2 is None or lat2 is None:
            return None
        
        import math
        
        R = 6371000  
        
        lat1_rad = math.radians(lat1)
        lon1_rad = math.radians(lon1)
        lat2_rad = math.radians(lat2)
        lon2_rad = math.radians(lon2)
        
        dlat = lat2_rad - lat1_rad
        dlon = lon2_rad - lon1_rad
        
        a = math.sin(dlat / 2) ** 2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2) ** 2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        
        distance = R * c
        return distance
    
    def _read_section_data(self, xlsx_path):
        """读取断面数据"""
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            
            points = []
            lon, lat = None, None
            deepest_lon, deepest_lat = None, None
            deepest_point = None
            
            found_data_start = False
            data_start_row = 13
            
            for row in range(1, min(ws.max_row + 1, 60)):
                col_b = ws.cell(row=row, column=2).value
                col_b_str = str(col_b).strip() if col_b else ''
                
                if '深泓' in col_b_str or '深泓点' in col_b_str:
                    deepest_lon = ws.cell(row=row, column=5).value
                    deepest_lat = ws.cell(row=row, column=6).value
                    self._log_action(f"    从特征点行找到深泓点经纬度: ({deepest_lon}, {deepest_lat})")
                
                if not found_data_start:
                    if '起点距' in str(ws.cell(row=row, column=3).value).strip():
                        found_data_start = True
                        data_start_row = row + 1
                        self._log_action(f"    找到数据起始行: {data_start_row}")
                    elif '垂线号' in col_b_str:
                        found_data_start = True
                        data_start_row = row + 1
            
            if not found_data_start:
                data_start_row = 13
            
            min_elevation = float('inf')
            deepest_x = None
            deepest_y = None
            
            for data_row in range(data_start_row, min(ws.max_row + 1, 200)):
                try:
                    x = ws.cell(row=data_row, column=3).value
                    y = ws.cell(row=data_row, column=4).value
                    
                    if x is None or y is None:
                        continue
                        
                    x = float(x)
                    y = float(y)
                    
                    if abs(x) < 10000 and abs(y) < 10000:
                        points.append((x, y))
                        
                        if y < min_elevation:
                            min_elevation = y
                            deepest_x = x
                            deepest_y = y
                except (ValueError, TypeError):
                    continue
            
            if deepest_x is not None:
                deepest_point = (deepest_x, deepest_y)
            
            for row in range(1, min(ws.max_row + 1, 20)):
                col_b = ws.cell(row=row, column=2).value
                col_b_str = str(col_b).strip() if col_b else ''
                if '经度' in col_b_str:
                    lon = ws.cell(row=row, column=4).value
                if '纬度' in col_b_str:
                    lat = ws.cell(row=row, column=4).value
            
            if deepest_lon is None and lon is not None:
                deepest_lon = lon
            if deepest_lat is None and lat is not None:
                deepest_lat = lat
            
            wb.close()
            self._log_action(f"    读取到 {len(points)} 个测点, 深泓点: {deepest_point}, 深泓点经纬度: ({deepest_lon}, {deepest_lat})")
            return points, lon, lat, deepest_point, deepest_lon, deepest_lat
        
        except Exception as e:
            self._log_action(f"读取断面数据失败 {xlsx_path}: {str(e)}")
            return [], None, None
    
    def _create_storage_data_template(self, template_path):
        """创建库容数据模板（包含公式）"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '库容计算'
        
        headers = [
            '断面名称', '深泓点经度', '深泓点纬度', '断面面积(m2)', 
            '桥名/断面', '断面序号', '桥顶高程', '经度', '纬度', 
            '断面面积(m2)', '棱柱高(m)', '体积V(m3)', '库容(m3)'
        ]
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        for col in ['L', 'M', 'N']:
            ws.column_dimensions[col].width = 18
        
        ws.cell(row=2, column=12).value = '0.000'
        ws.cell(row=2, column=13).value = '0.000'
        
        for row in range(3, 100):
            ws.cell(row=row, column=12).value = f'=IF(AND(J{row}<>"",J{row-1}<>"",K{row}<>"",K{row-1}<>"",K{row}-K{row-1}<>"",K{row}-K{row-1}>0),(J{row}+J{row-1})/2*(K{row}-K{row-1}),0)'
            ws.cell(row=row, column=13).value = f'=L{row-1}+L{row}'
        
        wb.save(template_path)
        self._log_action(f"创建库容数据模板: {template_path}")
    
    def _create_section_area_template(self, template_path):
        """创建断面面积计算模板（包含公式）"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '断面面积计算'
        
        headers = ['序号', '平距(m)', '高程(m)', '水深(m)', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        ws.cell(row=2, column=1, value='库容顶点高程(85):')
        ws.cell(row=2, column=2).number_format = '0.000'
        
        ws.cell(row=4, column=1, value='断面面积:')
        ws.cell(row=4, column=2).number_format = '0.000'
        
        ws.cell(row=2, column=4, value='计算说明:')
        ws.cell(row=3, column=4, value='水深=顶点高程-实际高程')
        ws.cell(row=4, column=4, value='面积=梯形面积累加')
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
        
        for row in range(6, 100):
            ws.cell(row=row, column=4).value = f'=MAX(0,$B$2-C{row})'
        
        ws.cell(row=102, column=2).value = '=SUMPRODUCT(ABS(B7:B99-B6:B98),(D6:D98+D7:D99)/2)'
        
        wb.save(template_path)
        self._log_action(f"创建断面面积计算模板: {template_path}")
    
    def _ensure_templates(self):
        """确保模板文件存在，不存在则创建"""
        if not os.path.exists(self.template_dir):
            os.makedirs(self.template_dir)
        
        if not os.path.exists(self.section_area_template_path):
            self._create_section_area_template(self.section_area_template_path)
        else:
            self._log_action(f"使用已存在的模板: {self.section_area_template_path}")
        
        if not os.path.exists(self.storage_data_template_path):
            self._create_storage_data_template(self.storage_data_template_path)
        else:
            self._log_action(f"使用已存在的模板: {self.storage_data_template_path}")
    
    def _get_mapping_data(self):
        """从对应表获取桥名和库容顶点高程（包含高程转换）"""
        mapping_file = os.path.join(self.template_dir, '对应表.xlsx')
        if not os.path.exists(mapping_file):
            mapping_file = os.path.join(self.output_dir, '对应表.xlsx')
        
        if not os.path.exists(mapping_file):
            self._log_action("错误: 对应表.xlsx 不存在")
            return {}
        
        bridge_data = {}
        
        try:
            wb = openpyxl.load_workbook(mapping_file, read_only=True)
            ws = None
            
            if 'Sheet6' in wb.sheetnames:
                ws = wb['Sheet6']
            elif len(wb.sheetnames) >= 6:
                ws = wb.worksheets[5]
            
            if ws is None:
                self._log_action("错误: 对应表缺少Sheet6")
                wb.close()
                return {}
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                section_id = ws.cell(row=2, column=col).value
                
                if cell_value:
                    bridge_name, elevation = self._extract_bridge_and_elevation(cell_value)
                    if bridge_name and elevation is not None:
                        if section_id:
                            subtract_value = self._find_subtract_value(section_id)
                            if subtract_value is not None:
                                elevation_85 = elevation - subtract_value
                                self._log_action(f"桥: {bridge_name}, 原高程: {elevation}, 差值: {subtract_value}, 85高程: {elevation_85}")
                            else:
                                elevation_85 = elevation
                                self._log_action(f"桥: {bridge_name}, 未找到高程差值，使用原值: {elevation}")
                        else:
                            elevation_85 = elevation
                            self._log_action(f"桥: {bridge_name}, 无断面编号，使用原值: {elevation}")
                        
                        bridge_data[bridge_name] = {
                            'elevation_85': elevation_85,
                            'section_id': section_id,
                            'original_elevation': elevation
                        }
            
            wb.close()
            self._log_action(f"从对应表读取到 {len(bridge_data)} 座桥的数据")
            return bridge_data
        
        except Exception as e:
            self._log_action(f"读取对应表失败: {str(e)}")
            return {}
    
    def _find_section_files(self, folder_path):
        """递归查找所有断面文件"""
        section_files = []
        
        for item in os.listdir(folder_path):
            item_path = os.path.join(folder_path, item)
            if os.path.isfile(item_path) and item.lower().endswith('.xlsx'):
                if not item.startswith('~$'):
                    section_files.append(item_path)
            elif os.path.isdir(item_path):
                section_files.extend(self._find_section_files(item_path))
        
        return section_files
    
    def _process_bridge_folder(self, bridge_folder_path, bridge_info):
        """处理单个桥文件夹"""
        bridge_name = bridge_folder_path.split(os.sep)[-1]
        vertex_elevation = bridge_info['elevation_85']
        
        self._log_action(f"处理桥: {bridge_name}")
        self._log_action(f"库容顶点高程(85): {vertex_elevation}")
        
        section_data = []
        
        section_files = self._find_section_files(bridge_folder_path)
        self._log_action(f"  找到 {len(section_files)} 个断面文件")
        
        for item_path in section_files:
            item = os.path.basename(item_path)
            result = self._read_section_data(item_path)
            
            if len(result) == 3:
                points, lon, lat = result
                deepest_point = None
                deepest_lon = None
                deepest_lat = None
            else:
                points, lon, lat, deepest_point, deepest_lon, deepest_lat = result
            
            if points:
                area = self._calculate_section_area(points, vertex_elevation)
                
                if area < 1.0:
                    avg_elevation = sum(p[1] for p in points) / len(points)
                    self._log_action(f"  警告: 断面 {item} 面积为{area:.2f}m2，顶点高程{vertex_elevation}m，测点平均高程{avg_elevation:.2f}m")
                    if vertex_elevation < avg_elevation - 0.1:
                        new_vertex = avg_elevation + 1.0
                        self._log_action(f"  自动调整顶点高程从{vertex_elevation}到{new_vertex}")
                        area = self._calculate_section_area(points, new_vertex)
                
                section_info = {
                    'name': os.path.splitext(item)[0],
                    'path': item_path,
                    'points': points,
                    'lon': lon,
                    'lat': lat,
                    'area': area,
                    'deepest': deepest_point,
                    'deepest_lon': deepest_lon,
                    'deepest_lat': deepest_lat
                }
                section_data.append(section_info)
                self._log_action(f"  断面 {section_info['name']}: 面积={area:.2f}m2")
        
        if len(section_data) < 2:
            self._log_action(f"  警告: 断面数量不足（{len(section_data)}个），无法计算库容")
            return None
        
        section_data.sort(key=lambda x: x['deepest'][1] if x['deepest'] else float('inf'))
        self._log_action(f"  断面已按深泓点高程排序（从下游到上游）")
        
        total_volume = 0.0
        volume_details = []
        
        for i in range(len(section_data) - 1):
            sec1 = section_data[i]
            sec2 = section_data[i + 1]
            
            if sec1['deepest_lon'] and sec1['deepest_lat'] and sec2['deepest_lon'] and sec2['deepest_lat']:
                distance = self._calculate_distance_by_latlon(sec1['deepest_lon'], sec1['deepest_lat'], sec2['deepest_lon'], sec2['deepest_lat'])
                distance_source = '经纬度'
            elif sec1['deepest'] and sec2['deepest']:
                distance = self._calculate_distance(sec1['deepest'], sec2['deepest'])
                distance_source = '平面坐标'
            else:
                self._log_action(f"  警告: 无法计算 {sec1['name']} - {sec2['name']} 的距离")
                continue
            
            if distance is None or distance < 0.1:
                self._log_action(f"  警告: 距离计算结果异常 {distance}m")
                continue
            
            avg_area = (sec1['area'] + sec2['area']) / 2
            volume = avg_area * distance
            
            total_volume += volume
            
            volume_details.append({
                'section1': sec1,
                'section2': sec2,
                'distance': distance,
                'distance_source': distance_source,
                'avg_area': avg_area,
                'volume': volume
            })
            
            self._log_action(f"  棱柱体 {i+1}: {sec1['name']} - {sec2['name']}, 距离={distance:.2f}m({distance_source}), 体积={volume:.2f}m3")
        
        self._log_action(f"  总库容: {total_volume:.2f}m3")
        
        return {
            'bridge_name': bridge_name,
            'vertex_elevation': vertex_elevation,
            'sections': section_data,
            'volume_details': volume_details,
            'total_volume': total_volume
        }
    
    def _export_to_excel(self, bridge_results):
        """导出计算结果到Excel"""
        os.makedirs(self.storage_output_dir, exist_ok=True)
        output_file = os.path.join(self.storage_output_dir, '库容数据.xlsx')
        
        if os.path.exists(self.storage_data_template_path):
            shutil.copy2(self.storage_data_template_path, output_file)
            self._log_action(f"使用模板创建: {output_file}")
        elif os.path.exists(output_file):
            pass
        else:
            self._create_storage_data_template(output_file)
        
        wb = openpyxl.load_workbook(output_file)
        
        for result in bridge_results:
            bridge_name = result['bridge_name']
            sections = result['sections']
            
            if bridge_name in wb.sheetnames:
                wb.remove(wb[bridge_name])
            
            template_ws = wb['库容计算']
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = bridge_name
            
            for row in range(3, new_ws.max_row + 1):
                for col in range(1, new_ws.max_column + 1):
                    new_ws.cell(row=row, column=col).value = None
            
            row_num = 2
            
            for i, sec in enumerate(sections):
                new_ws.cell(row=row_num, column=1).value = sec['name']
                new_ws.cell(row=row_num, column=2).value = sec['deepest_lon']
                new_ws.cell(row=row_num, column=3).value = sec['deepest_lat']
                new_ws.cell(row=row_num, column=4).value = sec['area']
                new_ws.cell(row=row_num, column=5).value = bridge_name
                new_ws.cell(row=row_num, column=6).value = row_num - 1
                new_ws.cell(row=row_num, column=7).value = result['vertex_elevation']
                new_ws.cell(row=row_num, column=8).value = sec['lon']
                new_ws.cell(row=row_num, column=9).value = sec['lat']
                new_ws.cell(row=row_num, column=10).value = sec['area']
                
                if row_num > 2:
                    prev_sec = sections[i - 1]
                    
                    if prev_sec['deepest_lon'] and prev_sec['deepest_lat'] and sec['deepest_lon'] and sec['deepest_lat']:
                        distance = self._calculate_distance_by_latlon(prev_sec['deepest_lon'], prev_sec['deepest_lat'], sec['deepest_lon'], sec['deepest_lat'])
                    elif prev_sec['deepest'] and sec['deepest']:
                        distance = self._calculate_distance(prev_sec['deepest'], sec['deepest'])
                    else:
                        distance = None
                    
                    if distance and distance > 0.1:
                        new_ws.cell(row=row_num, column=11).value = distance
                        
                        avg_area = (prev_sec['area'] + sec['area']) / 2
                        volume = avg_area * distance
                        new_ws.cell(row=row_num, column=12).value = volume
                
                row_num += 1
            
            total_volume_10k = result['total_volume'] / 10000
            new_ws.cell(row=row_num - 1, column=13).value = round(total_volume_10k, 4)
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                new_ws.column_dimensions[col].auto_size = True
        
        wb.save(output_file)
        self._log_action(f"导出结果到: {output_file}")
    
    def _generate_section_area_files(self, bridge_results, bridge_data):
        """生成断面面积计算文件"""
        for result in bridge_results:
            bridge_folder = os.path.join(self.storage_output_dir, '断面面积计算', result['bridge_name'])
            
            if not os.path.exists(bridge_folder):
                os.makedirs(bridge_folder)
            
            for sec in result['sections']:
                template_path = os.path.join(bridge_folder, f"{sec['name']}_面积计算.xlsx")
                
                if os.path.exists(self.section_area_template_path):
                    shutil.copy2(self.section_area_template_path, template_path)
                    wb = openpyxl.load_workbook(template_path)
                else:
                    wb = openpyxl.Workbook()
                
                ws = wb.active
                ws.title = '断面面积计算'
                
                ws.cell(row=2, column=2).value = result['vertex_elevation']
                
                row_num = 4
                for i, (x, y) in enumerate(sec['points'], 1):
                    ws.cell(row=row_num, column=1).value = i
                    ws.cell(row=row_num, column=2).value = x
                    ws.cell(row=row_num, column=3).value = y
                    if y is not None:
                        depth = max(0, result['vertex_elevation'] - y)
                        ws.cell(row=row_num, column=4).value = depth
                    row_num += 1
                
                ws.cell(row=row_num + 1, column=2).value = sec['area']
                
                wb.save(template_path)
                self._log_action(f"生成断面面积计算文件: {template_path}")
    
    def process_all(self, report_files=None, progress_callback=None):
        """执行完整流程"""
        self.results = {'success': [], 'failed': [], 'total': 0}
        self.log_text = ""
        self.progress_callback = progress_callback
        
        try:
            self._log_action("========== 库容计算开始 ==========")
            
            os.makedirs(self.storage_output_dir, exist_ok=True)
            self._log_action(f"输出目录: {self.storage_output_dir}")
            
            self._log_action("\n>> 步骤1: 确保模板文件存在")
            self._ensure_templates()
            
            self._log_action("\n>> 步骤2: 加载高程差值数据")
            self._load_ref_data()
            
            self._log_action("\n>> 步骤3: 读取对应表数据（含高程转换）")
            bridge_data = self._get_mapping_data()
            
            if not bridge_data:
                self._log_action("错误: 未读取到桥数据")
                self.results['failed'].append(('库容计算', '未读取到桥数据'))
                return self.results
            
            bridge_results = []
            
            self._log_action("\n>> 步骤4: 处理各桥数据")
            
            measure_table_dir = os.path.join(self.output_dir, '跨沟道路和桥涵', '测量表')
            if not os.path.exists(measure_table_dir):
                self._log_action(f"错误: 测量表目录不存在 {measure_table_dir}")
                self.results['failed'].append(('库容计算', '测量表目录不存在'))
                return self.results
            
            existing_folders = []
            for item in os.listdir(measure_table_dir):
                item_path = os.path.join(measure_table_dir, item)
                if os.path.isdir(item_path):
                    existing_folders.append(item)
            
            self._log_action(f"测量表目录下找到 {len(existing_folders)} 个文件夹")
            
            for bridge_name, info in bridge_data.items():
                if self._check_stop():
                    return self.results
                
                bridge_folder = None
                
                if bridge_name in existing_folders:
                    bridge_folder = os.path.join(measure_table_dir, bridge_name)
                else:
                    for folder_name in existing_folders:
                        if folder_name.endswith('_' + bridge_name) or bridge_name in folder_name:
                            bridge_folder = os.path.join(measure_table_dir, folder_name)
                            self._log_action(f"匹配到带编号的文件夹: {folder_name} -> {bridge_name}")
                            break
                
                if bridge_folder is None:
                    self._log_action(f"警告: 未找到桥 '{bridge_name}' 的文件夹")
                    continue
                
                if not os.path.exists(bridge_folder):
                    self._log_action(f"警告: 桥文件夹不存在 {bridge_folder}")
                    continue
                
                result = self._process_bridge_folder(bridge_folder, info)
                if result:
                    bridge_results.append(result)
            
            if not bridge_results:
                self._log_action("错误: 未处理到任何桥数据")
                self.results['failed'].append(('库容计算', '未处理到任何桥数据'))
                return self.results
            
            if self._check_stop():
                return self.results
            
            self._log_action("\n>> 步骤5: 生成断面面积计算文件")
            self._generate_section_area_files(bridge_results, bridge_data)
            
            if self._check_stop():
                return self.results
            
            self._log_action("\n>> 步骤6: 导出库容数据")
            self._export_to_excel(bridge_results)
            
            log_file_path = os.path.join(self.storage_output_dir, '库容计算日志.txt')
            with open(log_file_path, 'w', encoding='utf-8') as f:
                f.write(self.log_text)
            
            self.results['success'].append('库容计算完成')
            self._log_action("\n========== 库容计算完成 ==========")
            return self.results
        
        except Exception as e:
            self.results['failed'].append(('库容计算', str(e)))
            self._log_action(f"库容计算失败: {str(e)}")
            log_file_path = os.path.join(self.storage_output_dir, '库容计算日志.txt')
            with open(log_file_path, 'w', encoding='utf-8') as f:
                f.write(self.log_text)
            log_error("库容计算失败", e)
            return self.results
