import os
import openpyxl
from ..utils import log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR as OUTPUT_DIR, DEFAULT_TEMPLATE_DIR as TEMPLATE_DIR

class ReportTo85Converter:
    def __init__(self, output_dir=None, template_dir=None):
        self.results = {
            'success': [],
            'failed': [],
            'total': 0
        }
        self.ref_data = None
        self.output_dir = output_dir if output_dir else OUTPUT_DIR
        self.template_dir = template_dir if template_dir else TEMPLATE_DIR
    
    def load_ref_data(self):
        """加载对应表数据（Sheet3）"""
        try:
            table_path = os.path.join(self.template_dir, '对应表.xlsx')
            if not os.path.exists(table_path):
                table_path = os.path.join(self.template_dir, '对应表.xlsm')
                if not os.path.exists(table_path):
                    raise FileNotFoundError(f"对应表文件不存在: {table_path}")
            
            wb = openpyxl.load_workbook(table_path, data_only=True)
            if 'Sheet3' in wb.sheetnames:
                ws = wb['Sheet3']
            else:
                ws = wb.active
            
            self.ref_data = []
            for row in ws.iter_rows(values_only=True):
                self.ref_data.append(list(row))
            
            log_message("成功加载对应表数据")
            return True
        except Exception as e:
            log_error("加载对应表数据失败", e)
            return False
    
    def find_subtract_value(self, file_name):
        """根据文件名查找对应的高程差值"""
        if not self.ref_data or len(self.ref_data) < 2:
            return None
        
        search_name = os.path.splitext(file_name)[0]
        
        for row in self.ref_data[1:]:
            for idx, cell_value in enumerate(row):
                if cell_value and str(cell_value).strip() == search_name:
                    if idx < len(self.ref_data[0]):
                        subtract_value = self.ref_data[0][idx]
                        if isinstance(subtract_value, (int, float)):
                            return subtract_value
        return None
    
    def process_column(self, ws, col_letter, start_row, subtract_value):
        """处理指定列"""
        try:
            max_row = ws.max_row
            if max_row < start_row:
                return
            
            for row in range(start_row, max_row + 1):
                cell = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col_letter))
                if isinstance(cell.value, (int, float)) and isinstance(subtract_value, (int, float)):
                    cell.value = cell.value - subtract_value
        except Exception as e:
            log_error(f"处理列 {col_letter} 失败", e)
    
    def process_single_cell(self, ws, cell_address, subtract_value):
        """处理单个单元格"""
        try:
            cell = ws[cell_address]
            if isinstance(cell.value, (int, float)) and isinstance(subtract_value, (int, float)):
                cell.value = cell.value - subtract_value
        except Exception as e:
            log_error(f"处理单元格 {cell_address} 失败", e)
    
    def process_single_cell_if_value(self, ws, cell_address, subtract_value):
        """当单元格有值时处理"""
        try:
            cell = ws[cell_address]
            if cell.value is not None and isinstance(cell.value, (int, float)) and isinstance(subtract_value, (int, float)):
                cell.value = cell.value - subtract_value
        except Exception as e:
            log_error(f"处理单元格 {cell_address} (条件处理) 失败", e)
    
    def convert_to_85(self, report_file, progress_callback=None):
        """将成果表转换为85高程"""
        try:
            if not self.ref_data:
                if not self.load_ref_data():
                    raise ValueError("无法加载对应表")
            
            if not os.path.exists(report_file):
                raise FileNotFoundError(f"文件不存在: {report_file}")
            
            file_name = os.path.basename(report_file)
            search_name = os.path.splitext(file_name)[0]
            
            has_heng = '横' in search_name
            has_zong = '纵' in search_name
            
            if has_heng:
                process_type = '横断Z'
            elif has_zong:
                process_type = '纵Z'
            else:
                process_type = '纵Z' if 'Z' in search_name.upper() else '横断Z'
            
            subtract_value = self.find_subtract_value(file_name)
            if subtract_value is None:
                log_message(f"未找到文件 {file_name} 的高程差值，跳过")
                self.results['failed'].append((file_name, "未找到高程差值"))
                return False
            
            wb = openpyxl.load_workbook(report_file)
            ws = wb.active
            
            if process_type == '纵Z':
                self.process_column(ws, 'E', 11, subtract_value)
                self.process_column(ws, 'F', 11, subtract_value)
                self.process_single_cell(ws, 'D6', subtract_value)
            else:
                self.process_column(ws, 'D', 13, subtract_value)
                self.process_single_cell_if_value(ws, 'E7', subtract_value)
                self.process_single_cell_if_value(ws, 'E9', subtract_value)
                self.process_single_cell_if_value(ws, 'B10', subtract_value)
            
            output_path = report_file
            wb.save(output_path)
            
            self.results['success'].append(file_name)
            log_message(f"成功转换85高程: {file_name}")
            
            if progress_callback:
                progress_callback(f"已转换85高程: {file_name}")
            
            return True
        except Exception as e:
            file_name = os.path.basename(report_file)
            self.results['failed'].append((file_name, str(e)))
            log_error(f"转换85高程失败: {file_name}", e)
            return False
    
    def process_all(self, report_files, progress_callback=None):
        """批量转换所有成果表"""
        self.results = {'success': [], 'failed': [], 'total': len(report_files)}
        completed_count = 0
        
        if not self.ref_data:
            if not self.load_ref_data():
                return self.results
        
        for report_file in report_files:
            self.convert_to_85(report_file, None)
            
            completed_count += 1
            if progress_callback:
                progress_callback(f"已转换85高程: {os.path.basename(report_file)}")
                progress_callback(f"已完成 {completed_count}/{len(report_files)}")
        
        if progress_callback and self.results['failed']:
            progress_callback(f"\n===== 失败详情 =====")
            for item in self.results['failed']:
                progress_callback(f"失败: {item[0]} - {item[1]}")
        
        return self.results