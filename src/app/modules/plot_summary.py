import os
import openpyxl
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from ..utils import log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR as OUTPUT_DIR

class PlotSummary:
    def __init__(self, output_dir=None):
        self.results = {
            'success': [],
            'failed': [],
            'total': 0
        }
        self.output_dir = output_dir if output_dir else OUTPUT_DIR
    
    def _is_report_file(self, file_name):
        """判断是否为成果表文件（只处理包含'成果表'关键词的文件）"""
        if '成果表' in file_name:
            return True
        return False
    
    def _get_data_range(self, file_name):
        """根据文件名确定数据提取范围"""
        if '纵断面' in file_name:
            return (11, 7, 8)
        else:
            return (13, 5, 6)
    
    def _extract_data(self, report_file):
        """提取单个成果表的数据"""
        file_name = os.path.basename(report_file)
        
        if not self._is_report_file(file_name):
            return None
        
        ext = file_name.split('.')[-1].lower() if '.' in file_name else ''
        if ext not in ['xlsx', 'xls', 'xlsm']:
            return None
        
        try:
            wb = openpyxl.load_workbook(report_file, data_only=True, read_only=True)
            ws = wb.active
            
            start_row, start_col, end_col = self._get_data_range(file_name)
            file_base_name = os.path.splitext(file_name)[0]
            
            last_row = ws.max_row
            for row in range(ws.max_row, 1, -1):
                cell_value = ws.cell(row=row, column=end_col).value
                if cell_value is not None and str(cell_value).strip() != '':
                    last_row = row
                    break
            
            if last_row < start_row:
                last_row = start_row
            
            data_lines = []
            for row in range(start_row, last_row + 1):
                row_data = file_base_name
                is_empty = True
                
                for col in range(start_col, end_col + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        cell_str = str(cell_value).strip()
                        row_data += ',' + cell_str
                        if cell_str:
                            is_empty = False
                
                if not is_empty:
                    data_lines.append(row_data)
            
            wb.close()
            return (file_name, data_lines)
        except Exception as e:
            return (file_name, None)
    
    def export_to_txt(self, report_files, progress_callback=None, max_workers=4):
        """将所有成果表的坐标数据导出到一个TXT文件（多线程并行处理）"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.output_dir, f'成果汇总_{timestamp}.txt')
            
            filtered_files = []
            for report_file in report_files:
                file_name = os.path.basename(report_file)
                if self._is_report_file(file_name):
                    ext = file_name.split('.')[-1].lower() if '.' in file_name else ''
                    if ext in ['xlsx', 'xls', 'xlsm']:
                        filtered_files.append(report_file)
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = [executor.submit(self._extract_data, rf) for rf in filtered_files]
                
                results = []
                completed_count = 0
                for future in as_completed(futures):
                    result = future.result()
                    completed_count += 1
                    if result:
                        results.append(result)
                    
                    if progress_callback:
                        progress_callback(f"已读取 {completed_count}/{len(filtered_files)}")
            
            with open(output_path, 'w', encoding='utf-8') as f:
                for file_name, data_lines in results:
                    if data_lines:
                        for line in data_lines:
                            f.write(line + '\n')
                        f.write('\n')
                        self.results['success'].append(file_name)
                        log_message(f"成功处理成果表: {file_name}")
                    else:
                        self.results['failed'].append((file_name, "提取数据失败"))
                        log_error(f"处理成果表失败: {file_name}", "提取数据失败")
            
            log_message(f"导出完成，共处理 {len(self.results['success'])} 个成果表，输出文件: {output_path}")
            return True
            
        except Exception as e:
            log_error("导出TXT失败", e)
            return False
    
    def process_all(self, report_files, progress_callback=None):
        """批量导出TXT（所有成果表数据合并到一个文件）"""
        self.results = {'success': [], 'failed': [], 'total': len(report_files)}
        return self.export_to_txt(report_files, progress_callback)