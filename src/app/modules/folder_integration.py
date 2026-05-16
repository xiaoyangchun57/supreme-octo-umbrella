import os
import shutil
import openpyxl
from ..utils import log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR as OUTPUT_DIR, DEFAULT_TEMPLATE_DIR as TEMPLATE_DIR

class FolderIntegrator:
    def __init__(self, output_dir=None, template_dir=None):
        self.results = {
            "success": [],
            "failed": [],
            "total": 0
        }
        self.output_dir = output_dir if output_dir else OUTPUT_DIR
        self.template_dir = template_dir if template_dir else TEMPLATE_DIR
        self.log_text = ""
    
    def _log_action(self, message):
        """记录日志"""
        self.log_text += message + "\n"
        log_message(message)
    
    def _clean_filename(self, input_name):
        """清理文件名中的非法字符"""
        bad_chars = ["\\", "/", "?", "*", ":", "|", "<", ">", "\""]
        for char in bad_chars:
            input_name = input_name.replace(char, "_")
        return input_name
    
    def _clean_sheetname(self, input_name):
        """清理工作表名称中的非法字符"""
        bad_chars = ["\\", "/", "?", "*", "[", "]", ":"]
        for char in bad_chars:
            input_name = input_name.replace(char, "")
        return input_name
    
    def _search_and_copy_files(self, root_path, dest_folder_path, file_dict):
        """搜索并复制文件"""
        for item in os.listdir(root_path):
            item_path = os.path.join(root_path, item)
            
            if os.path.isfile(item_path):
                filename = os.path.basename(item_path)
                if filename != "操作记录.txt" and filename != "对应表.XLSX" and filename != "对应表.xlsx":
                    base_name = os.path.splitext(filename)[0]
                    if base_name in file_dict:
                        dest_file_path = os.path.join(dest_folder_path, filename)
                        if not os.path.exists(dest_file_path):
                            shutil.copy2(item_path, dest_file_path)
                            self._log_action(f"复制文件: {item_path} -> {dest_file_path}")
                        else:
                            self._log_action(f"已存在: {filename} 在 {dest_folder_path}")
            
            elif os.path.isdir(item_path):
                if item_path != dest_folder_path:
                    self._search_and_copy_files(item_path, dest_folder_path, file_dict)
    
    def organize_files(self, root_path):
        """模型8：整理文件夹"""
        try:
            structure_path = os.path.join(root_path, "对应表.XLSX")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(root_path, "对应表.xlsx")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(root_path, "对应表.xlsm")
            
            if not os.path.exists(structure_path):
                structure_path = os.path.join(self.template_dir, "对应表.XLSX")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(self.template_dir, "对应表.xlsx")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(self.template_dir, "对应表.xlsm")
            
            if not os.path.exists(structure_path):
                self._log_action(f"错误: 对应表不存在")
                raise FileNotFoundError(f"对应表不存在")
            
            self._log_action(f"找到对应表: {structure_path}")
            wb = openpyxl.load_workbook(structure_path, data_only=True)
            
            self._log_action(f"===== 开始整理 {root_path} =====")
            self._log_action(f"工作表数量: {len(wb.sheetnames)}")
            
            for sheet_num in [4, 5]:
                if sheet_num > len(wb.sheetnames):
                    self._log_action(f"警告: 未找到Sheet{sheet_num}，跳过")
                    continue
                
                ws = wb.worksheets[sheet_num - 1]
                self._log_action(f"正在处理Sheet{sheet_num}")
                
                main_folder_a1 = str(ws.cell(row=1, column=1).value or "").strip()
                sub_folder_a2 = str(ws.cell(row=2, column=1).value or "").strip()
                
                self._log_action(f"A1(一级文件夹): '{main_folder_a1}'")
                self._log_action(f"A2(二级文件夹): '{sub_folder_a2}'")
                
                if not main_folder_a1 or not sub_folder_a2:
                    self._log_action(f"警告: Sheet{sheet_num}的A1/A2为空，跳过")
                    continue
                
                main_path = os.path.join(root_path, main_folder_a1)
                if not os.path.exists(main_path):
                    os.makedirs(main_path)
                    self._log_action(f"创建一级文件夹: {main_path}")
                
                sub_path = os.path.join(main_path, sub_folder_a2)
                if not os.path.exists(sub_path):
                    os.makedirs(sub_path)
                    self._log_action(f"创建二级文件夹: {sub_path}")
                
                last_col = 1
                for col in range(1, 100):
                    if ws.cell(row=3, column=col).value is None:
                        break
                    last_col = col
                
                self._log_action(f"总列数: {last_col}")
                
                for col_index in range(1, last_col + 1):
                    sub_folder_name = str(ws.cell(row=3, column=col_index).value or "").strip()
                    
                    if not sub_folder_name:
                        self._log_action(f"警告: 第{col_index}列三级文件夹名为空，跳过")
                        continue
                    
                    third_path = os.path.join(sub_path, sub_folder_name)
                    if not os.path.exists(third_path):
                        os.makedirs(third_path)
                        self._log_action(f"创建三级文件夹: {third_path}")
                    
                    file_dict = {}
                    last_row = 3
                    for row in range(4, 1000):
                        cell_value = ws.cell(row=row, column=col_index).value
                        if cell_value is None:
                            break
                        filename = str(cell_value).strip()
                        if filename:
                            file_dict[filename] = filename
                    
                    if file_dict:
                        self._log_action(f"第{col_index}列: 复制 {len(file_dict)} 个文件到 {sub_folder_name}")
                        self._search_and_copy_files(root_path, third_path, file_dict)
            
            wb.close()
            
            log_file_path = os.path.join(root_path, "操作记录.txt")
            with open(log_file_path, "w", encoding="utf-8") as f:
                f.write(self.log_text)
            
            self._log_action(f"===== 整理完成 =====")
            return True
            
        except Exception as e:
            self.results["failed"].append(("文件夹整理", str(e)))
            log_error("文件夹整理失败", e)
            return False
    
    def _find_file(self, filename, search_folder):
        """搜索文件"""
        search_path = os.path.join(search_folder, filename + ".xlsx")
        if os.path.exists(search_path):
            return search_path
        
        for item in os.listdir(search_folder):
            item_path = os.path.join(search_folder, item)
            if os.path.isdir(item_path):
                result = self._find_file(filename, item_path)
                if result:
                    return result
        return None
    
    def _copy_worksheet(self, source_path, target_workbook, sheet_name):
        """复制工作表（包括图表）"""
        try:
            source_wb = openpyxl.load_workbook(source_path, read_only=False, data_only=False)
            
            # 获取源文件的第一个sheet
            source_ws = source_wb.worksheets[0] if source_wb.worksheets else None
            
            if source_ws:
                # 复制整个工作表到目标工作簿
                new_ws = target_workbook.create_sheet(title=self._clean_sheetname(sheet_name))
                
                # 复制所有单元格值和格式
                for row in source_ws.iter_rows():
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = cell.font.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.number_format = cell.number_format
                            new_cell.protection = cell.protection.copy()
                            new_cell.alignment = cell.alignment.copy()
                
                # 复制图表
                if hasattr(source_ws, "_charts") and source_ws._charts:
                    for chart in source_ws._charts:
                        new_ws.add_chart(chart)
            
            source_wb.close()
        except Exception as e:
            self._log_action(f"复制失败: {source_path} -> {sheet_name} | 错误: {str(e)}")
    
    def merge_files(self, root_path):
        """模型9：合并文件"""
        try:
            structure_path = os.path.join(root_path, "对应表.XLSX")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(root_path, "对应表.xlsx")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(root_path, "对应表.xlsm")
            
            if not os.path.exists(structure_path):
                structure_path = os.path.join(self.template_dir, "对应表.XLSX")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(self.template_dir, "对应表.xlsx")
            if not os.path.exists(structure_path):
                structure_path = os.path.join(self.template_dir, "对应表.xlsm")
            if not os.path.exists(structure_path):
                raise FileNotFoundError(f"对应表不存在: {structure_path}")
            
            self._log_action(f"===== 开始合并 {root_path} =====")
            
            wb = openpyxl.load_workbook(structure_path, data_only=True)
            
            for sheet_num in [4, 5]:
                if sheet_num > len(wb.sheetnames):
                    self._log_action(f"警告: 未找到Sheet{sheet_num}，跳过")
                    continue
                
                ws = wb.worksheets[sheet_num - 1]
                
                # A1: 一级文件夹名
                base_folder_name = str(ws.cell(row=1, column=1).value or "").strip()
                # B2: 合并后存放的文件夹名
                new_folder_name = str(ws.cell(row=2, column=2).value or "").strip()
                
                if not base_folder_name or not new_folder_name:
                    self._log_action(f"警告: Sheet{sheet_num}的A1/B2为空，跳过")
                    continue
                
                # 检查一级文件夹是否存在
                base_folder_path = os.path.join(root_path, base_folder_name)
                if not os.path.exists(base_folder_path):
                    self._log_action(f"警告: 目标文件夹不存在: {base_folder_path}，跳过")
                    continue
                
                # 创建新的合并文件夹
                new_folder_path = os.path.join(base_folder_path, new_folder_name)
                if not os.path.exists(new_folder_path):
                    os.makedirs(new_folder_path)
                    self._log_action(f"创建合并文件夹: {new_folder_path}")
                
                # 获取所有列
                last_col = 1
                for col in range(1, 100):
                    if ws.cell(row=3, column=col).value is None:
                        break
                    last_col = col
                
                self._log_action(f"Sheet{sheet_num}: 处理 {last_col} 列")
                
                # 处理每一列
                for col_index in range(1, last_col + 1):
                    header_value = str(ws.cell(row=3, column=col_index).value or "").strip()
                    
                    if not header_value:
                        continue
                    
                    # 获取目标文件名（取下划线前的部分）
                    parts = header_value.split("_")
                    if parts:
                        target_file_name = self._clean_filename(parts[0])
                    else:
                        target_file_name = self._clean_filename(header_value)
                    
                    if not target_file_name:
                        continue
                    
                    target_path = os.path.join(new_folder_path, target_file_name + ".xlsx")
                    if os.path.exists(target_path):
                        os.remove(target_path)
                    
                    # 创建目标工作簿
                    target_wb = openpyxl.Workbook()
                    
                    # 获取文件列表（从第4行开始）
                    file_count = 0
                    for row in range(4, 1000):
                        cell_value = ws.cell(row=row, column=col_index).value
                        if cell_value is None:
                            break
                        
                        file_value = str(cell_value).strip()
                        if file_value:
                            # 取下划线后的部分作为搜索文件名
                            file_parts = file_value.split("_")
                            if len(file_parts) >= 2:
                                search_name = file_parts[1].strip()
                            else:
                                search_name = file_value.strip()
                            
                            if search_name:
                                source_path = self._find_file(search_name, root_path)
                                if source_path:
                                    self._log_action(f"复制工作表: {os.path.basename(source_path)}")
                                    self._copy_worksheet(source_path, target_wb, search_name)
                                    file_count += 1
                    
                    # 删除默认的Sheet1
                    if "Sheet" in target_wb.sheetnames:
                        target_wb.remove(target_wb["Sheet"])
                    
                    # 保存目标工作簿
                    if len(target_wb.sheetnames) > 0:
                        target_wb.save(target_path)
                        self._log_action(f"已创建: {target_file_name}.xlsx ({file_count} 个工作表)")
                    else:
                        self._log_action(f"跳过: {target_file_name}.xlsx (无工作表)")
                    
                    target_wb.close()
            
            wb.close()
            
            log_file_path = os.path.join(os.path.expanduser("~"), "Desktop", "文件合并记录.txt")
            with open(log_file_path, "w", encoding="utf-8") as f:
                f.write(self.log_text)
            
            self._log_action("===== 合并完成 =====")
            return True
            
        except Exception as e:
            self.results["failed"].append(("文件合并", str(e)))
            log_error("文件合并失败", e)
            return False
    
    def process_all(self, root_path, progress_callback=None):
        """执行完整的整合流程"""
        self.results = {"success": [], "failed": [], "total": 2}
        self.log_text = ""
        
        self.organize_files(root_path)
        self.merge_files(root_path)
        
        return self.results
