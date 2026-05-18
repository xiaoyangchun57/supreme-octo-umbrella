import os
import shutil
import openpyxl
import zipfile
import re
import gc
import time
from threading import Lock
from concurrent.futures import ThreadPoolExecutor, as_completed
from ..utils import log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR as OUTPUT_DIR, DEFAULT_TEMPLATE_DIR as TEMPLATE_DIR

class AutoPlotter:
    def __init__(self, output_dir=None, template_dir=None):
        self.results = {
            'success': [],
            'failed': [],
            'total': 0
        }
        self.dict_sheet1 = {}
        self.dict_sheet2 = {}
        self.output_dir = output_dir if output_dir else OUTPUT_DIR
        self.template_dir = template_dir if template_dir else TEMPLATE_DIR
        self._file_locks = {}
        self._lock = Lock()
    
    def _load_mapping_dict(self):
        """从对应表加载映射字典"""
        table_path = os.path.join(self.template_dir, '对应表.xlsx')
        if not os.path.exists(table_path):
            table_path = os.path.join(self.template_dir, '对应表.xlsm')
            if not os.path.exists(table_path):
                log_message("对应表不存在，跳过映射加载")
                return
        
        try:
            wb = openpyxl.load_workbook(table_path, data_only=True)
            
            if 'Sheet1' in wb.sheetnames:
                self._fill_dictionary_from_sheet(wb['Sheet1'], self.dict_sheet1, 11)
            
            if 'Sheet2' in wb.sheetnames:
                self._fill_dictionary_from_sheet(wb['Sheet2'], self.dict_sheet2, 8)
            
            wb.close()
            wb = None
            log_message(f"成功加载映射字典，Sheet1: {len(self.dict_sheet1)}条，Sheet2: {len(self.dict_sheet2)}条")
        except Exception as e:
            log_error("加载映射字典失败", e)
    
    def _fill_dictionary_from_sheet(self, ws, target_dict, col_index):
        """从工作表填充字典（提取文件名最后一段作为key）"""
        last_row = ws.max_row
        for i in range(1, last_row + 1):
            cell_value = ws.cell(row=i, column=col_index).value
            if cell_value:
                cell_str = str(cell_value).strip()
                parts = cell_str.split('_')
                if len(parts) > 0:
                    stake = parts[-1]
                    if '.' in stake:
                        stake = os.path.splitext(stake)[0]
                    f_value = ws.cell(row=i, column=6).value
                    target_dict[stake] = f_value
    
    def _find_last_row(self, ws, cols):
        """查找指定列中的最后一行数据"""
        col_list = cols.split(',')
        last_row = 0
        for col in col_list:
            col = col.strip()
            col_idx = openpyxl.utils.column_index_from_string(col)
            for row in range(ws.max_row, 0, -1):
                if ws.cell(row=row, column=col_idx).value is not None:
                    if row > last_row:
                        last_row = row
                    break
        return last_row
    
    def _transfer_data(self, src_path, dest_path):
        """传输数据（严格按照VB代码逻辑）"""
        wb_src = None
        wb_dest = None
        try:
            wb_src = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
            src_sheet = wb_src.active
            
            wb_dest = openpyxl.load_workbook(dest_path)
            dest_sheet = wb_dest.active
            
            src_file_name = os.path.basename(src_path)
            is_profile = '纵断面' in src_file_name
            
            if is_profile:
                last_row = self._find_last_row(src_sheet, 'C,E')
                if last_row < 11:
                    raise ValueError("纵断面数据不足，需要至少到C11:E11")
                
                col_c = []
                col_e = []
                for row in range(11, last_row + 1):
                    c_val = src_sheet.cell(row=row, column=3).value
                    e_val = src_sheet.cell(row=row, column=5).value
                    col_c.append(c_val)
                    col_e.append(e_val)
                
                cumulative = 0.0
                num_rows = len(col_c)
                src_data = []
                for i in range(num_rows):
                    if isinstance(col_c[i], (int, float)):
                        cumulative += float(col_c[i])
                    src_data.append([cumulative, col_e[i]])
            else:
                last_row = self._find_last_row(src_sheet, 'C,D')
                if last_row < 13:
                    raise ValueError("横断面数据不足，需要至少到C13:D13")
                
                src_data = []
                for row in range(13, last_row + 1):
                    c_val = src_sheet.cell(row=row, column=3).value
                    d_val = src_sheet.cell(row=row, column=4).value
                    src_data.append([c_val, d_val])
            
            dest_sheet['A3'] = "平距/m"
            dest_sheet['B3'] = "高程/m"
            
            for row in range(4, 201):
                dest_sheet.cell(row=row, column=1, value=None)
                dest_sheet.cell(row=row, column=2, value=None)
            
            for i, row_data in enumerate(src_data):
                dest_sheet.cell(row=4 + i, column=1, value=row_data[0])
                dest_sheet.cell(row=4 + i, column=2, value=row_data[1])
            
            wb_dest.save(dest_path)
            wb_dest.close()
            wb_dest = None
            wb_src.close()
            wb_src = None
            
            return True
        except Exception as e:
            log_error("数据传输失败", e)
            return False
        finally:
            try:
                if wb_src:
                    wb_src.close()
                    wb_src = None
            except:
                pass
            try:
                if wb_dest:
                    wb_dest.close()
                    wb_dest = None
            except:
                pass
    
    def _modify_output_file(self, file_path):
        """修改输出文件（填写表头信息）"""
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path)
            sht = wb.active
            
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            
            sht['B2'] = file_name
            sht.title = file_name[:31]
            sht.sheet_properties.tabColor = "FFC000"
            
            value_to_fill = ""
            if file_name in self.dict_sheet1:
                value_to_fill = self.dict_sheet1[file_name]
            elif file_name in self.dict_sheet2:
                value_to_fill = self.dict_sheet2[file_name]
            
            if value_to_fill:
                sht['A2'] = value_to_fill
            
            title_text = sht['B2'].value
            if title_text:
                title_text = str(title_text).strip()
            else:
                title_text = file_name
            
            for chart in sht._charts:
                chart.title = title_text
            
            wb.save(file_path)
            wb.close()
            wb = None
            
            import zipfile
            import re
            
            temp_path = file_path + '.tmp'
            
            with zipfile.ZipFile(file_path, 'r') as z:
                content = z.read('xl/charts/chart1.xml').decode('utf-8')
            
            content = content.replace("'1'!", f"'{file_name}'!")
            
            content = re.sub(r"<c:numCache>.*?</c:numCache>", "<c:numCache><c:ptCount val=\"0\"/></c:numCache>", content, flags=re.DOTALL)
            
            with zipfile.ZipFile(temp_path, 'w') as z_out:
                with zipfile.ZipFile(file_path, 'r') as z_in:
                    for name in z_in.namelist():
                        if name == 'xl/charts/chart1.xml':
                            z_out.writestr(name, content.encode('utf-8'))
                        else:
                            z_out.writestr(name, z_in.read(name))
            
            gc.collect()
            time.sleep(0.3)
            
            max_retries = 3
            retry_delay = 0.2
            
            for attempt in range(max_retries):
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    os.rename(temp_path, file_path)
                    break
                except (PermissionError, OSError) as e:
                    if attempt < max_retries - 1:
                        log_message(f"文件操作重试 {attempt + 1}/{max_retries}: {str(e)}")
                        gc.collect()
                        time.sleep(retry_delay)
                    else:
                        raise
            
            return True
        except Exception as e:
            log_error("修改输出文件失败", e)
            return False
        finally:
            try:
                if wb:
                    wb.close()
                    wb = None
            except:
                pass
    
    def generate_plot(self, report_file, progress_callback=None):
        """根据成果表生成图形（严格按照VB代码逻辑）"""
        try:
            if not os.path.exists(report_file):
                raise FileNotFoundError(f"文件不存在: {report_file}")
            
            template_path = os.path.join(self.template_dir, '成图模板.xlsx')
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"成图模板不存在: {template_path}")
            
            if not self.dict_sheet1:
                self._load_mapping_dict()
            
            file_name = os.path.basename(report_file)
            base_name = os.path.splitext(file_name)[0]
            
            parts = base_name.split('_')
            if len(parts) > 1:
                output_base_name = parts[-1]
            else:
                output_base_name = base_name
            
            output_name = f"{output_base_name}.xlsx"
            output_path = os.path.join(self.output_dir, output_name)
            
            gc.collect()
            time.sleep(0.3)
            
            max_retries = 3
            retry_delay = 0.2
            
            for attempt in range(max_retries):
                try:
                    if os.path.exists(output_path):
                        try:
                            os.remove(output_path)
                        except (PermissionError, OSError) as e:
                            if attempt < max_retries - 1:
                                log_message(f"删除旧文件重试 {attempt + 1}/{max_retries}")
                                gc.collect()
                                time.sleep(retry_delay)
                                continue
                            else:
                                raise
                    shutil.copy(template_path, output_path)
                    break
                except Exception as e:
                    if attempt < max_retries - 1:
                        gc.collect()
                        time.sleep(retry_delay)
                    else:
                        raise
            
            if not self._transfer_data(report_file, output_path):
                self.results['failed'].append((file_name, "数据传输失败"))
                return False
            
            if not self._modify_output_file(output_path):
                self.results['failed'].append((file_name, "修改输出文件失败"))
                return False
            
            self.results['success'].append(output_name)
            log_message(f"成功生成图形: {file_name} -> {output_name}")
            
            if progress_callback:
                progress_callback(f"已生成图形: {file_name}")
            
            return True
        except Exception as e:
            file_name = os.path.basename(report_file)
            self.results['failed'].append((file_name, str(e)))
            log_error(f"生成图形失败: {file_name}", e)
            return False
    
    def _generate_plot_thread(self, report_file):
        """单线程生成图形（供多线程调用）"""
        max_retries = 2
        retry_delay = 0.5
        
        for attempt in range(max_retries):
            try:
                gc.collect()
                
                if not os.path.exists(report_file):
                    return False, os.path.basename(report_file), "文件不存在"
                
                template_path = os.path.join(self.template_dir, '成图模板.xlsx')
                if not os.path.exists(template_path):
                    return False, os.path.basename(report_file), "成图模板不存在"
                
                file_name = os.path.basename(report_file)
                base_name = os.path.splitext(file_name)[0]
                
                parts = base_name.split('_')
                if len(parts) > 1:
                    output_base_name = parts[-1]
                else:
                    output_base_name = base_name
                
                output_name = f"{output_base_name}.xlsx"
                output_path = os.path.join(self.output_dir, output_name)
                
                gc.collect()
                time.sleep(0.3)
                
                if os.path.exists(output_path):
                    try:
                        os.remove(output_path)
                        gc.collect()
                        time.sleep(0.2)
                    except (PermissionError, OSError) as e:
                        if attempt < max_retries - 1:
                            log_message(f"删除旧文件重试 {attempt + 1}/{max_retries}: {file_name}")
                            gc.collect()
                            time.sleep(retry_delay)
                            continue
                        else:
                            return False, os.path.basename(report_file), f"文件操作失败: {str(e)}"
                
                shutil.copy(template_path, output_path)
                gc.collect()
                time.sleep(0.2)
                
                if not self._transfer_data(report_file, output_path):
                    if attempt < max_retries - 1:
                        log_message(f"数据传输失败，重试 {attempt + 1}/{max_retries}: {file_name}")
                        gc.collect()
                        time.sleep(retry_delay)
                        continue
                    return False, file_name, "数据传输失败"
                
                if not self._modify_output_file(output_path):
                    if attempt < max_retries - 1:
                        log_message(f"修改输出文件失败，重试 {attempt + 1}/{max_retries}: {file_name}")
                        gc.collect()
                        time.sleep(retry_delay)
                        continue
                    return False, file_name, "修改输出文件失败"
                
                return True, file_name, output_name
            except (PermissionError, OSError) as e:
                if attempt < max_retries - 1:
                    log_message(f"文件操作冲突，重试 {attempt + 1}/{max_retries}: {os.path.basename(report_file)}")
                    gc.collect()
                    time.sleep(retry_delay)
                else:
                    return False, os.path.basename(report_file), f"文件操作失败: {str(e)}"
            except Exception as e:
                return False, os.path.basename(report_file), str(e)
    
    def process_all(self, report_files, progress_callback=None, max_workers=4):
        """批量生成图形（使用合理并发线程）"""
        self.results = {'success': [], 'failed': [], 'total': len(report_files)}
        
        if not self.dict_sheet1:
            self._load_mapping_dict()
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []
            for report_file in report_files:
                future = executor.submit(self._generate_plot_thread, report_file)
                futures.append(future)
            
            completed_count = 0
            for future in as_completed(futures):
                success, file_name, output_name = future.result()
                completed_count += 1
                
                if success:
                    self.results['success'].append(output_name)
                    log_message(f"成功生成图形: {file_name} -> {output_name}")
                else:
                    self.results['failed'].append((file_name, output_name))
                    log_error(f"生成图形失败: {file_name}", output_name)
                
                if progress_callback:
                    progress_callback(f"已完成 {completed_count}/{len(report_files)}")
        
        if progress_callback and self.results['failed']:
            progress_callback(f"\n===== 失败详情 =====")
            for item in self.results['failed']:
                progress_callback(f"失败: {item[0]} - {item[1]}")
        
        return self.results
