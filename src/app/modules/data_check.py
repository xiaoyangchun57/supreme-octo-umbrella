import os
import openpyxl
from ..utils import log_message, log_error
from ..config import DEFAULT_OUTPUT_DIR

class DataChecker:
    def __init__(self, output_dir=None):
        self.results = {
            'success': [],
            'failed': [],
            'warnings': []
        }
        self.output_dir = output_dir if output_dir else DEFAULT_OUTPUT_DIR
    
    def _is_target_folder(self, file_path):
        """判断是否为目标文件夹（防治对象、跨沟道路、桥涵）"""
        folder_path = os.path.dirname(file_path)
        folder_name = os.path.basename(folder_path).lower()
        return '防治对象' in folder_name or '跨沟道路' in folder_name or '桥涵' in folder_name
    
    def _is_cross_section_report(self, xlsx_file):
        """判断是否为横断面成果表"""
        if not self._is_target_folder(xlsx_file):
            return False
        file_name = os.path.basename(xlsx_file).lower()
        return '横断面' in file_name or 'zbc' in file_name
    
    def _find_header_row(self, ws):
        """查找表头行（包含'垂线号'的行）"""
        for row_idx in range(1, min(ws.max_row + 1, 30)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and '垂线号' in str(cell_value):
                return row_idx
        return None
    
    def _find_data_end_row(self, ws, start_row):
        """查找数据结束行（连续空行的开始）"""
        empty_count = 0
        max_empty = 3
        for row_idx in range(start_row + 1, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 8)]
            if all(v is None or str(v).strip() == '' for v in row_values):
                empty_count += 1
                if empty_count >= max_empty:
                    return row_idx - max_empty
            else:
                empty_count = 0
        return ws.max_row
    
    def check_report_file(self, xlsx_file, progress_callback=None):
        """检查横断面成果表"""
        try:
            file_name = os.path.basename(xlsx_file)
            
            if not self._is_cross_section_report(xlsx_file):
                if progress_callback:
                    progress_callback(f"跳过（非横断面）: {file_name}")
                return True
            
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            errors = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                header_row = self._find_header_row(ws)
                if header_row is None:
                    errors.append(f"Sheet '{sheet_name}': 未找到表头行（垂线号）")
                    continue
                
                data_start_row = header_row + 1
                data_end_row = self._find_data_end_row(ws, data_start_row)
                
                found_features = {
                    '左堤顶': False,
                    '右堤顶': False,
                    '深泓点': False
                }
                roughness_issues = []
                distance_issues = []
                
                prev_distance = None
                
                for row_idx in range(data_start_row, data_end_row + 1):
                    feature_desc = ws.cell(row=row_idx, column=2).value
                    distance = ws.cell(row=row_idx, column=3).value
                    roughness = ws.cell(row=row_idx, column=7).value
                    
                    if feature_desc:
                        feature_str = str(feature_desc).strip()
                        if '左堤顶' in feature_str:
                            found_features['左堤顶'] = True
                        if '右堤顶' in feature_str:
                            found_features['右堤顶'] = True
                        if '深泓点' in feature_str:
                            found_features['深泓点'] = True
                    
                    if roughness is None or str(roughness).strip() == '':
                        roughness_issues.append(f"第{row_idx}行糙率为空")
                    
                    if distance is not None:
                        try:
                            dist_val = float(distance)
                            if prev_distance is not None and dist_val < prev_distance:
                                distance_issues.append(f"第{row_idx}行起点距({dist_val})小于前一行({prev_distance})")
                            prev_distance = dist_val
                        except (ValueError, TypeError):
                            pass
                
                missing_features = [k for k, v in found_features.items() if not v]
                if missing_features:
                    errors.append(f"Sheet '{sheet_name}': 缺少特征点 - {', '.join(missing_features)}")
                
                if roughness_issues:
                    errors.append(f"Sheet '{sheet_name}': 糙率填写不完整（{len(roughness_issues)}处）")
                    for issue in roughness_issues[:3]:
                        errors.append(f"  {issue}")
                    if len(roughness_issues) > 3:
                        errors.append(f"  ... 等共{len(roughness_issues)}处")
                
                if distance_issues:
                    errors.append(f"Sheet '{sheet_name}': 起点距不递增（{len(distance_issues)}处）")
                    for issue in distance_issues[:3]:
                        errors.append(f"  {issue}")
                    if len(distance_issues) > 3:
                        errors.append(f"  ... 等共{len(distance_issues)}处")
            
            wb.close()
            
            if errors:
                self.results['warnings'].append((file_name, errors))
                log_message(f"成果表检查警告: {file_name} - {'; '.join(errors[:2])}")
            else:
                self.results['success'].append(file_name)
                log_message(f"成果表检查通过: {file_name}")
            
            if progress_callback:
                progress_callback(f"已检查: {file_name}")
            
            return len(errors) == 0
        except Exception as e:
            file_name = os.path.basename(xlsx_file)
            self.results['failed'].append((file_name, str(e)))
            log_error(f"成果表检查失败: {file_name}", e)
            return False
    
    def find_empty_sections_in_report(self, xlsx_files, progress_callback=None):
        """查找成果表中的空白断面"""
        try:
            empty_files = []
            
            for xlsx_file in xlsx_files:
                if not self._is_cross_section_report(xlsx_file):
                    continue
                    
                try:
                    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                    
                    has_data = False
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        header_row = self._find_header_row(ws)
                        if header_row:
                            for row_idx in range(header_row + 1, min(ws.max_row + 1, 50)):
                                if ws.cell(row=row_idx, column=2).value is not None:
                                    has_data = True
                                    break
                        if has_data:
                            break
                    
                    wb.close()
                    
                    if not has_data:
                        empty_files.append(os.path.basename(xlsx_file))
                except Exception:
                    empty_files.append(os.path.basename(xlsx_file))
            
            if empty_files:
                self.results['warnings'].append(('空白成果表', empty_files))
                log_message(f"发现空白成果表: {', '.join(empty_files)}")
            
            if progress_callback:
                progress_callback(f"空白成果表查找完成，发现{len(empty_files)}个空白成果表")
            
            return empty_files
        except Exception as e:
            log_error("查找空白成果表失败", e)
            return []
    
    def check_depth_consistency(self, xlsx_file, progress_callback=None):
        """检查成果表深度数据一致性（别名方法，保持兼容性）"""
        return self.check_report_file(xlsx_file, progress_callback)
    
    def generate_report(self):
        """生成检查报告"""
        try:
            report_content = []
            report_content.append("=" * 60)
            report_content.append("横断面成果表数据检查报告")
            report_content.append("=" * 60)
            report_content.append("")
            report_content.append("检查内容：")
            report_content.append("  1. B列中断面特征点（左堤顶、右堤顶、深泓点）")
            report_content.append("  2. G列中糙率填写完整性")
            report_content.append("  3. C列中起点距递增性")
            report_content.append("")
            report_content.append(f"检查通过: {len(self.results['success'])}")
            report_content.append(f"检查失败: {len(self.results['failed'])}")
            report_content.append(f"存在问题: {len(self.results['warnings'])}")
            report_content.append("")
            report_content.append("-" * 60)
            
            if self.results['success']:
                report_content.append("通过文件:")
                for item in self.results['success']:
                    report_content.append(f"  ✓ {item}")
                report_content.append("")
            
            if self.results['failed']:
                report_content.append("失败文件:")
                for item, error in self.results['failed']:
                    report_content.append(f"  ✗ {item}: {error}")
                report_content.append("")
            
            if self.results['warnings']:
                report_content.append("存在问题:")
                for item, warnings in self.results['warnings']:
                    report_content.append(f"  ⚠ {item}:")
                    if isinstance(warnings, list):
                        for warning in warnings:
                            report_content.append(f"    - {warning}")
                    else:
                        report_content.append(f"    - {warnings}")
                    report_content.append("")
            
            os.makedirs(self.output_dir, exist_ok=True)
            report_path = os.path.join(self.output_dir, '横断面成果表检查报告.txt')
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(report_content))
            
            log_message("成功生成横断面成果表检查报告")
            return report_path
        except Exception as e:
            log_error("生成检查报告失败", e)
            return None
